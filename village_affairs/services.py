import base64
import json
import math
import uuid
from datetime import date, datetime, timedelta
from io import StringIO
from pathlib import Path
from django.apps import apps
from django.core.files.base import ContentFile
from django.core.management import call_command

import openpyxl
import xlrd
from django.conf import settings
from django.core.paginator import Paginator
from django.db import IntegrityError, OperationalError, ProgrammingError, transaction
from django.db.models import Count, Q
from django.utils import timezone

from .models import (
    DatabaseBackupRecord,
    Household,
    MigrantWorker,
    OperationLog,
    Resident,
    ResidentImportBatch,
    RiskCheck,
    UISettings,
    VillageGroup,
)

IMPORTABLE_EXTENSIONS = {'.xlsx', '.xls'}
DATABASE_BACKUP_EXTENSIONS = {'.json'}

SYSTEM_FIELDS = [
    {'key': 'full_name', 'label': '居民姓名', 'required': True},
    {'key': 'gender', 'label': '性别', 'required': False},
    {'key': 'identity_number', 'label': '身份证号码', 'required': True},
    {'key': 'birth_date', 'label': '出生日期', 'required': False},
    {'key': 'ethnicity', 'label': '民族', 'required': False},
    {'key': 'phone', 'label': '联系电话', 'required': False},
    {'key': 'head_name', 'label': '户主姓名', 'required': False},
    {'key': 'head_identity_number', 'label': '户主身份证号码', 'required': False},
    {'key': 'relation_to_head', 'label': '与户主关系', 'required': False},
    {'key': 'village_group', 'label': '村组', 'required': False},
    {'key': 'address', 'label': '家庭地址', 'required': False},
    {'key': 'household_type', 'label': '户属性', 'required': False},
    {'key': 'account_type', 'label': '户口类型', 'required': False},
    {'key': 'grid_name', 'label': '所属网格', 'required': False},
    {'key': 'political_status', 'label': '政治面貌', 'required': False},
    {'key': 'marital_status', 'label': '婚姻状况', 'required': False},
    {'key': 'health_status', 'label': '健康状态', 'required': False},
    {'key': 'residency_status', 'label': '居住状态', 'required': False},
    {'key': 'education_level', 'label': '文化程度', 'required': False},
    {'key': 'occupation', 'label': '职业', 'required': False},
    {'key': 'bank_account', 'label': '银行账号', 'required': False},
    {'key': 'bank_name', 'label': '开户行', 'required': False},
    {'key': 'military_status', 'label': '兵役状况', 'required': False},
    {'key': 'status', 'label': '状态', 'required': False},
    {'key': 'household_no', 'label': '户编号', 'required': False},
    {'key': 'housing_type', 'label': '住房类型', 'required': False},
]

FIELD_LABELS = {item['key']: item['label'] for item in SYSTEM_FIELDS}

OPTION_FIELDS = {
    'gender': ['男', '女'],
    'ethnicity': ['汉族', '回族', '满族', '蒙古族', '藏族', '维吾尔族'],
    'relation_to_head': ['户主', '配偶', '子女', '父母', '兄弟姐妹', '其他'],
    'household_type': ['普通户', '低保户', '贫困户', '五保户'],
    'account_type': ['农业户口', '非农业户口'],
    'grid_name': ['网格一', '网格二', '网格三'],
    'political_status': ['群众', '共产党员', '共青团员', '民主党派'],
    'marital_status': ['未婚', '已婚', '离婚', '丧偶'],
    'health_status': ['健康', '一般', '残疾', '重病'],
    'residency_status': ['常住', '流动', '外出务工', '已迁出'],
    'education_level': ['文盲', '小学', '初中', '高中', '大专', '本科', '硕士', '博士'],
    'bank_name': ['中国农业银行', '中国工商银行', '中国建设银行', '中国银行', '邮政储蓄银行'],
    'military_status': ['未服兵役', '现役', '退役'],
    'status': ['正常', '停用'],
    'housing_type': ['自建房', '租房', '公租房', '商品房'],
    'village_group': ['一组', '二组', '三组', '四组', '五组'],
}

DETAIL_EXPORT_COLUMNS = [
    ('居民姓名', 'full_name'),
    ('性别', 'gender'),
    ('身份证号码', 'identity_number'),
    ('出生日期', 'birth_date'),
    ('年龄', 'age'),
    ('民族', 'ethnicity'),
    ('联系电话', 'phone'),
    ('户主姓名', 'head_name'),
    ('户主身份证号码', 'head_identity_number'),
    ('与户主关系', 'relation_to_head'),
    ('村组', 'village_group'),
    ('家庭地址', 'address'),
    ('户属性', 'household_type'),
    ('所属网格', 'grid_name'),
    ('政治面貌', 'political_status'),
    ('婚姻状况', 'marital_status'),
    ('健康状态', 'health_status'),
    ('居住状态', 'residency_status'),
    ('文化程度', 'education_level'),
    ('职业', 'occupation'),
    ('银行账号', 'bank_account'),
    ('开户行', 'bank_name'),
    ('兵役状况', 'military_status'),
    ('状态', 'status'),
]

HOUSEHOLD_EXPORT_COLUMNS = [
    ('户主姓名', 'head_name'),
    ('性别', 'head_gender'),
    ('村组', 'village_group'),
    ('家庭地址', 'address'),
    ('户主身份证号码', 'head_identity_number'),
    ('家庭人数', 'member_count'),
    ('户属性', 'household_type'),
    ('所属网格', 'grid_name'),
]

DEFAULT_UI_SETTINGS = {
    'system_title': '村务管理系统',
    'logo_mode': UISettings.LOGO_MODE_BOTH,
    'logo_text': '村务管理系统',
    'logo_image': '',
    'favicon': '',
    'village_overview': '请在系统设置-界面设置中维护村情概况，用于展示本村历史沿革、产业特色、公共服务设施和治理成果。',
    'village_image': '',
}

UI_SETTINGS_FIELDS = (
    'system_title',
    'logo_mode',
    'logo_text',
    'logo_image',
    'favicon',
    'village_overview',
    'village_image',
)

DEFAULT_OPERATOR_NAME = '系统管理员'
DEFAULT_OPERATION_LOG_RETENTION_DAYS = 90


def serialize_ui_settings(item):
    return {
        'systemTitle': item.system_title,
        'logoMode': item.logo_mode,
        'logoText': item.logo_text,
        'logoImage': item.logo_image.url if item.logo_image else '',
        'favicon': item.favicon.url if item.favicon else '',
        'villageOverview': item.village_overview,
        'villageImage': item.village_image.url if item.village_image else '',
    }


def get_ui_settings_instance():
    defaults = {
        'village_overview': DEFAULT_UI_SETTINGS['village_overview'],
        **{key: value for key, value in DEFAULT_UI_SETTINGS.items() if key != 'village_overview'},
    }
    item, _ = UISettings.objects.get_or_create(pk=1, defaults=defaults)
    return item


def get_ui_settings_payload():
    return serialize_ui_settings(get_ui_settings_instance())


def normalize_ui_settings_payload(payload):
    logo_mode = str(payload.get('logoMode') or DEFAULT_UI_SETTINGS['logo_mode']).strip()
    if logo_mode not in {
        UISettings.LOGO_MODE_IMAGE,
        UISettings.LOGO_MODE_TEXT,
        UISettings.LOGO_MODE_BOTH,
    }:
        logo_mode = DEFAULT_UI_SETTINGS['logo_mode']

    system_title = str(payload.get('systemTitle') or '').strip() or DEFAULT_UI_SETTINGS['system_title']
    logo_text = str(payload.get('logoText') or '').strip()
    village_overview = str(payload.get('villageOverview') or '').strip() or DEFAULT_UI_SETTINGS['village_overview']

    return {
        'system_title': system_title[:128],
        'logo_mode': logo_mode,
        'logo_text': logo_text[:128],
        'logo_image': str(payload.get('logoImage') or '').strip(),
        'favicon': str(payload.get('favicon') or '').strip(),
        'village_overview': village_overview[:2000],
        'village_image': str(payload.get('villageImage') or '').strip(),
    }


def update_ui_settings(payload):
    item = get_ui_settings_instance()
    normalized = normalize_ui_settings_payload(payload)
    
    # Handle scalar fields
    for field in ['system_title', 'logo_mode', 'logo_text', 'village_overview']:
        setattr(item, field, normalized[field])
        
    # Handle base64 image fields
    for field in ['logo_image', 'favicon', 'village_image']:
        data = normalized[field]
        if not data:
            getattr(item, field).delete(save=False)
        elif data.startswith('data:image'):
            try:
                format_part, imgstr = data.split(';base64,')
                ext = format_part.split('/')[-1].split(';')[0]
                if ext == 'svg+xml':
                    ext = 'svg'
                elif ext == 'jpeg':
                    ext = 'jpg'
                getattr(item, field).delete(save=False)
                getattr(item, field).save(f"{uuid.uuid4().hex}.{ext}", ContentFile(base64.b64decode(imgstr)), save=False)
            except Exception as e:
                pass # If parsing fails, ignore the image update

    item.save()
    return serialize_ui_settings(item)


def get_request_operator_name(request):
    if hasattr(request, 'session') and request.session.get('user_name'):
        return request.session['user_name']
    user = getattr(request, 'user', None)
    if user is not None and getattr(user, 'is_authenticated', False):
        username = user.get_username().strip()
        if username:
            return username
    header_name = str(request.headers.get('X-Operator-Name') or request.META.get('HTTP_X_OPERATOR_NAME') or '').strip()
    return header_name or DEFAULT_OPERATOR_NAME


def get_request_ip_address(request):
    forwarded_for = str(request.META.get('HTTP_X_FORWARDED_FOR') or '').strip()
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    return str(request.META.get('REMOTE_ADDR') or '').strip()


def record_operation_log(
    *,
    module,
    action,
    summary,
    operator=DEFAULT_OPERATOR_NAME,
    target='',
    result=OperationLog.RESULT_SUCCESS,
    detail='',
    ip_address='',
):
    try:
        return OperationLog.objects.create(
            module=str(module or '').strip()[:64] or '未分类模块',
            action=str(action or '').strip()[:32] or '未知操作',
            operator=str(operator or DEFAULT_OPERATOR_NAME).strip()[:64] or DEFAULT_OPERATOR_NAME,
            summary=str(summary or '').strip()[:255] or '未命名操作',
            target=str(target or '').strip()[:128],
            result=result if result in {OperationLog.RESULT_SUCCESS, OperationLog.RESULT_FAILED} else OperationLog.RESULT_SUCCESS,
            detail=str(detail or '').strip()[:4000],
            ip_address=str(ip_address or '').strip()[:64],
        )
    except (OperationalError, ProgrammingError, IntegrityError):
        return None


def serialize_operation_log(item, index=None):
    return {
        'id': item.id,
        'seq': index,
        'time': timezone.localtime(item.created_at).strftime('%Y-%m-%d %H:%M:%S'),
        'operator': item.operator,
        'module': item.module,
        'action': item.action,
        'summary': item.summary,
        'target': item.target,
        'result': item.result,
        'detail': item.detail,
        'ip_address': item.ip_address,
    }


def build_operation_log_queryset(params):
    queryset = OperationLog.objects.all()
    keyword = serialize_scalar(params.get('keyword'))
    module = serialize_scalar(params.get('module'))
    result = serialize_scalar(params.get('result'))
    created_from = parse_date_value(params.get('created_from'))
    created_to = parse_date_value(params.get('created_to'))

    if keyword:
        queryset = queryset.filter(
            Q(operator__icontains=keyword)
            | Q(summary__icontains=keyword)
            | Q(target__icontains=keyword)
            | Q(detail__icontains=keyword)
        )
    if module:
        queryset = queryset.filter(module=module)
    if result:
        queryset = queryset.filter(result=result)
    if created_from:
        queryset = queryset.filter(created_at__date__gte=created_from)
    if created_to:
        queryset = queryset.filter(created_at__date__lte=created_to)
    return queryset


def list_operation_logs(params):
    page_number = int(params.get('page', 1) or 1)
    page_size = int(params.get('page_size', 20) or 20)
    queryset = build_operation_log_queryset(params)
    paginator, page = paginate_queryset(queryset, page_number, page_size)
    start_index = (page.number - 1) * page_size + 1
    items = [serialize_operation_log(item, start_index + offset) for offset, item in enumerate(page.object_list)]
    modules = list(OperationLog.objects.order_by('module').values_list('module', flat=True).distinct())
    return {
        'items': items,
        'pagination': {
            'page': page.number,
            'page_size': page_size,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
        },
        'filter_options': {
            'modules': modules,
            'results': [OperationLog.RESULT_SUCCESS, OperationLog.RESULT_FAILED],
        },
    }


def cleanup_expired_operation_logs(retention_days=DEFAULT_OPERATION_LOG_RETENTION_DAYS):
    days = int(retention_days or DEFAULT_OPERATION_LOG_RETENTION_DAYS)
    if days <= 0:
        raise ValueError('保留天数必须大于 0。')
    threshold = timezone.now() - timedelta(days=days)
    deleted_count, _ = OperationLog.objects.filter(created_at__lt=threshold).delete()
    return {
        'message': f'已清理 {deleted_count} 条超过 {days} 天的日志。',
        'deleted_count': deleted_count,
        'retention_days': days,
    }


def get_database_backup_dir():
    backup_dir = settings.BASE_DIR / 'runtime' / 'database_backups'
    backup_dir.mkdir(parents=True, exist_ok=True)
    return backup_dir


def count_database_records():
    total = 0
    for model in apps.get_models():
        meta = model._meta
        if meta.proxy or meta.auto_created or not meta.managed:
            continue
        total += model.objects.count()
    return total


def format_file_size(file_size):
    size = int(file_size or 0)
    if size < 1024:
        return f'{size} B'
    if size < 1024 * 1024:
        return f'{size / 1024:.2f} KB'
    if size < 1024 * 1024 * 1024:
        return f'{size / (1024 * 1024):.2f} MB'
    return f'{size / (1024 * 1024 * 1024):.2f} GB'


def serialize_database_backup_record(item, index=None):
    file_path = Path(item.file_path)
    try:
        relative_path = file_path.relative_to(settings.BASE_DIR).as_posix()
    except ValueError:
        relative_path = item.file_path
    return {
        'id': item.id,
        'seq': index,
        'file_name': item.file_name,
        'file_path': item.file_path,
        'relative_path': relative_path,
        'file_size': item.file_size,
        'file_size_display': format_file_size(item.file_size),
        'record_count': item.record_count,
        'created_by': item.created_by,
        'status': item.status,
        'error_message': item.error_message,
        'created_at': timezone.localtime(item.created_at).strftime('%Y-%m-%d %H:%M:%S'),
        'file_exists': file_path.exists(),
    }


def list_database_backups():
    queryset = DatabaseBackupRecord.objects.all()
    items = [serialize_database_backup_record(item, index + 1) for index, item in enumerate(queryset)]
    return {
        'backup_dir': str(get_database_backup_dir()),
        'items': items,
        'summary': {
            'total': queryset.count(),
            'success_count': queryset.filter(status=DatabaseBackupRecord.STATUS_SUCCESS).count(),
            'failed_count': queryset.filter(status=DatabaseBackupRecord.STATUS_FAILED).count(),
        },
    }


def create_database_backup(operator=DEFAULT_OPERATOR_NAME):
    backup_dir = get_database_backup_dir()
    file_name = f"database-backup-{timezone.localtime(timezone.now()).strftime('%Y%m%d-%H%M%S-%f')}.json"
    file_path = backup_dir / file_name
    try:
        stream = StringIO()
        call_command('dumpdata', format='json', indent=2, stdout=stream)
        payload = stream.getvalue()
        json.loads(payload)
        file_path.write_text(payload, encoding='utf-8')
        item = DatabaseBackupRecord.objects.create(
            file_name=file_name,
            file_path=str(file_path),
            file_size=file_path.stat().st_size,
            record_count=count_database_records(),
            created_by=str(operator or DEFAULT_OPERATOR_NAME).strip()[:64] or DEFAULT_OPERATOR_NAME,
            status=DatabaseBackupRecord.STATUS_SUCCESS,
        )
        return serialize_database_backup_record(item)
    except Exception as exc:
        if file_path.exists():
            file_path.unlink()
        DatabaseBackupRecord.objects.create(
            file_name=file_name,
            file_path=str(file_path),
            created_by=str(operator or DEFAULT_OPERATOR_NAME).strip()[:64] or DEFAULT_OPERATOR_NAME,
            status=DatabaseBackupRecord.STATUS_FAILED,
            error_message=str(exc)[:4000],
        )
        raise ValueError(f'数据库备份失败：{exc}') from exc


def validate_database_backup_filename(filename):
    suffix = Path(filename or '').suffix.lower()
    if suffix not in DATABASE_BACKUP_EXTENSIONS:
        raise ValueError('仅支持导入 .json 格式的数据库备份文件。')
    return suffix


def load_database_backup_fixture(file_path):
    try:
        payload = json.loads(Path(file_path).read_text(encoding='utf-8'))
    except UnicodeDecodeError as exc:
        raise ValueError('备份文件编码无效，请上传 UTF-8 编码的 JSON 文件。') from exc
    except json.JSONDecodeError as exc:
        raise ValueError('备份文件不是有效的 JSON 数据。') from exc

    if not isinstance(payload, list):
        raise ValueError('备份文件格式不正确，应为 dumpdata 导出的 JSON 数组。')

    for index, item in enumerate(payload, start=1):
        if not isinstance(item, dict):
            raise ValueError(f'备份文件第 {index} 条记录格式不正确。')
        if 'model' not in item or 'fields' not in item:
            raise ValueError(f'备份文件第 {index} 条记录缺少必要字段。')

    return payload


def restore_database_from_backup_file(file_path, operator=DEFAULT_OPERATOR_NAME):
    fixture_items = load_database_backup_fixture(file_path)
    safety_backup = create_database_backup(operator)
    try:
        call_command('flush', interactive=False, verbosity=0, inhibit_post_migrate=True)
        call_command('loaddata', str(Path(file_path)), verbosity=0, ignorenonexistent=True)
    except Exception as exc:
        safety_file_name = safety_backup.get('file_name') if isinstance(safety_backup, dict) else ''
        safety_hint = f'；恢复前自动备份文件为 {safety_file_name}' if safety_file_name else ''
        raise ValueError(f'数据库恢复失败：{exc}{safety_hint}') from exc

    return {
        'message': '数据库恢复完成，当前数据已按备份文件内容全量覆盖。',
        'source_file_name': Path(file_path).name,
        'restored_records': len(fixture_items),
        'safety_backup_file_name': safety_backup.get('file_name') if isinstance(safety_backup, dict) else '',
        'safety_backup_path': safety_backup.get('file_path') if isinstance(safety_backup, dict) else '',
    }


def restore_database_backup(record_id, operator=DEFAULT_OPERATOR_NAME):
    item, file_path = get_database_backup_file(record_id)
    payload = restore_database_from_backup_file(file_path, operator)
    payload['source_file_name'] = item.file_name
    return payload


def restore_uploaded_database_backup(uploaded_file, operator=DEFAULT_OPERATOR_NAME):
    if uploaded_file is None:
        raise ValueError('请先选择要导入的数据库备份文件。')

    validate_database_backup_filename(getattr(uploaded_file, 'name', ''))
    temp_dir = get_database_backup_dir() / '_restore_uploads'
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_path = temp_dir / f"restore-upload-{timezone.localtime(timezone.now()).strftime('%Y%m%d-%H%M%S-%f')}-{uuid.uuid4().hex[:8]}.json"

    try:
        with temp_path.open('wb') as stream:
            for chunk in uploaded_file.chunks():
                stream.write(chunk)
        payload = restore_database_from_backup_file(temp_path, operator)
        payload['source_file_name'] = getattr(uploaded_file, 'name', temp_path.name)
        return payload
    finally:
        if temp_path.exists():
            temp_path.unlink()


def get_database_backup_file(record_id):
    item = DatabaseBackupRecord.objects.get(id=record_id)
    file_path = Path(item.file_path)
    if not file_path.exists():
        raise FileNotFoundError('备份文件不存在，可能已被手动删除。')
    return item, file_path


@transaction.atomic
def delete_database_backup(record_id):
    item = DatabaseBackupRecord.objects.get(id=record_id)
    file_path = Path(item.file_path)
    if file_path.exists():
        file_path.unlink()
    file_name = item.file_name
    item.delete()
    return {'message': f'备份文件 {file_name} 已删除。'}


def clear_all_database_data(operator):
    """清除数据库中所有业务数据。
    
    先创建安全备份，然后执行 flush 清空所有表数据。
    此操作不跳过 Django 的 migration 表，确保数据完全清空。
    
    Args:
        operator: 操作人姓名（用于备份记录，但不记录操作日志）
    
    Returns:
        dict: {'message': ..., 'backup': {...}}
    """
    from django.db import connection
    from django.core.management import call_command as _call_command
    
    # 1) 创建安全备份
    backup = create_database_backup(operator)
    
    # 2) 清空所有表数据
    _call_command('flush', '--no-input', verbosity=0)
    
    # 3) 重新运行迁移以确保所有表结构完整
    _call_command('migrate', '--run-syncdb', verbosity=0)
    _call_command('migrate', verbosity=0)
    
    return {
        'message': '所有数据已成功清除。系统已恢复为初始状态。',
        'backup': backup,
    }


def get_import_dir():
    import_dir = settings.BASE_DIR / 'runtime' / 'resident_imports'
    import_dir.mkdir(parents=True, exist_ok=True)
    return import_dir


def validate_excel_filename(filename):
    suffix = Path(filename or '').suffix.lower()
    if suffix not in IMPORTABLE_EXTENSIONS:
        raise ValueError('仅支持上传 .xlsx 和 .xls 格式的 Excel 文件。')
    return suffix


def serialize_scalar(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def load_excel_rows(file_path):
    suffix = Path(file_path).suffix.lower()
    if suffix == '.xlsx':
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not raw_rows:
            return [], []
        headers = [serialize_scalar(cell) for cell in raw_rows[0]]
        rows = []
        for raw_row in raw_rows[1:]:
            values = [serialize_scalar(cell) for cell in raw_row]
            if any(value for value in values):
                rows.append(dict(zip(headers, values)))
        return headers, rows

    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        return [], []

    def xls_value(cell, rowx, colx):
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                dt = xlrd.xldate_as_datetime(cell.value, workbook.datemode)
                return dt.date().isoformat()
            except (TypeError, ValueError):
                return ''
        if cell.ctype == xlrd.XL_CELL_NUMBER:
            value = float(cell.value)
            if value.is_integer():
                return str(int(value))
            return str(value)
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return '是' if cell.value else '否'
        return serialize_scalar(cell.value)

    headers = [xls_value(sheet.cell(0, col), 0, col) for col in range(sheet.ncols)]
    rows = []
    for row_idx in range(1, sheet.nrows):
        values = [xls_value(sheet.cell(row_idx, col), row_idx, col) for col in range(sheet.ncols)]
        if any(value for value in values):
            rows.append(dict(zip(headers, values)))
    return headers, rows


def suggest_mapping(headers):
    suggestions = {}
    for field in SYSTEM_FIELDS:
        label = field['label']
        for header in headers:
            if header == label or header.replace(' ', '') == label.replace(' ', ''):
                suggestions[field['key']] = header
                break
    return suggestions


def create_import_batch(uploaded_file):
    validate_excel_filename(uploaded_file.name)
    batch_id = uuid.uuid4()
    target_path = get_import_dir() / f'{batch_id}{Path(uploaded_file.name).suffix.lower()}'
    with target_path.open('wb') as output:
        for chunk in uploaded_file.chunks():
            output.write(chunk)

    headers, rows = load_excel_rows(target_path)
    batch = ResidentImportBatch.objects.create(
        id=batch_id,
        original_filename=uploaded_file.name,
        file_path=str(target_path),
        source_headers=headers,
        total_rows=len(rows),
    )
    return batch, headers, rows


def normalize_identity_number(value):
    return serialize_scalar(value).replace(' ', '').upper()


def parse_date_value(value):
    text = serialize_scalar(value)
    if not text:
        return None
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def derive_birth_date_from_identity(identity_number):
    identity_number = normalize_identity_number(identity_number)
    if len(identity_number) >= 14 and identity_number[:17].isdigit():
        birthday = identity_number[6:14]
        try:
            return datetime.strptime(birthday, '%Y%m%d').date()
        except ValueError:
            return None
    return None


def normalize_mapping(mapping):
    normalized = {}
    for key, value in (mapping or {}).items():
        if key in FIELD_LABELS and value:
            normalized[key] = str(value).strip()
    return normalized


def map_row_data(row, mapping):
    return {
        field_key: serialize_scalar(row.get(header))
        for field_key, header in mapping.items()
        if header
    }


def normalize_resident_payload(mapped_row):
    identity_number = normalize_identity_number(mapped_row.get('identity_number'))
    head_identity_number = normalize_identity_number(mapped_row.get('head_identity_number'))
    full_name = serialize_scalar(mapped_row.get('full_name'))
    relation_to_head = serialize_scalar(mapped_row.get('relation_to_head'))
    is_household_head = relation_to_head in {'户主', '本人'} or (
        identity_number and head_identity_number and identity_number == head_identity_number
    )
    birth_date = parse_date_value(mapped_row.get('birth_date')) or derive_birth_date_from_identity(identity_number)

    normalized = {
        'full_name': full_name,
        'gender': serialize_scalar(mapped_row.get('gender')),
        'identity_number': identity_number,
        'birth_date': birth_date,
        'ethnicity': serialize_scalar(mapped_row.get('ethnicity')),
        'phone': serialize_scalar(mapped_row.get('phone')),
        'head_name': serialize_scalar(mapped_row.get('head_name')),
        'head_identity_number': head_identity_number,
        'relation_to_head': relation_to_head,
        'village_group': serialize_scalar(mapped_row.get('village_group')),
        'address': serialize_scalar(mapped_row.get('address')),
        'household_type': serialize_scalar(mapped_row.get('household_type')),
        'account_type': serialize_scalar(mapped_row.get('account_type')),
        'grid_name': serialize_scalar(mapped_row.get('grid_name')),
        'political_status': serialize_scalar(mapped_row.get('political_status')),
        'marital_status': serialize_scalar(mapped_row.get('marital_status')),
        'health_status': serialize_scalar(mapped_row.get('health_status')),
        'residency_status': serialize_scalar(mapped_row.get('residency_status')),
        'education_level': serialize_scalar(mapped_row.get('education_level')),
        'occupation': serialize_scalar(mapped_row.get('occupation')),
        'bank_account': serialize_scalar(mapped_row.get('bank_account')),
        'bank_name': serialize_scalar(mapped_row.get('bank_name')),
        'military_status': serialize_scalar(mapped_row.get('military_status')),
        'status': serialize_scalar(mapped_row.get('status')) or '正常',
        'household_no': serialize_scalar(mapped_row.get('household_no')),
        'housing_type': serialize_scalar(mapped_row.get('housing_type')),
        'is_household_head': is_household_head,
    }
    if not normalized['head_name'] and is_household_head:
        normalized['head_name'] = full_name
    if not normalized['head_identity_number'] and is_household_head:
        normalized['head_identity_number'] = identity_number
    return normalized


def validate_identity_number(identity_number):
    identity_number = identity_number.strip()
    if not identity_number:
        return False, '身份证号码不能为空'
    if len(identity_number) not in (15, 18):
        return False, '身份证号码长度必须为15或18位'
    if len(identity_number) == 18:
        if not identity_number[:17].isdigit():
            return False, '身份证号码前17位必须为数字'
        check_code = identity_number[17].upper()
        if check_code not in '0123456789X':
            return False, '身份证号码校验码必须为数字或X'
    else:
        if not identity_number.isdigit():
            return False, '身份证号码必须为数字'
    return True, ''


def validate_phone(phone):
    phone = phone.strip()
    if not phone:
        return True, ''
    phone = phone.replace(' ', '').replace('-', '')
    if phone.startswith('+86'):
        phone = phone[3:]
    if len(phone) != 11:
        return False, '手机号必须为11位'
    if not phone.isdigit():
        return False, '手机号必须为数字'
    if not phone.startswith(('13', '14', '15', '16', '17', '18', '19')):
        return False, '手机号格式不正确'
    return True, ''


def validate_date(date_str):
    if not date_str:
        return True, ''
    try:
        dt = parse_date_value(date_str)
        if dt is None:
            return False, '日期格式不正确，支持 YYYY-MM-DD、YYYY/MM/DD、YYYY.MM.DD、YYYYMMDD 格式'
        return True, ''
    except Exception as e:
        return False, '日期格式不正确'


def validate_option_field(value, field_name, valid_options):
    if not value:
        return True, ''
    value = value.strip()
    if value and value not in valid_options:
        return False, f'{field_name}必须为以下选项之一: {", ".join(valid_options)}'
    return True, ''


def validate_normalized_row(normalized):
    errors = []
    
    if not normalized['full_name']:
        errors.append('缺少居民姓名')
    
    id_valid, id_msg = validate_identity_number(normalized['identity_number'])
    if not id_valid:
        errors.append(id_msg)
    
    if normalized['phone']:
        phone_valid, phone_msg = validate_phone(normalized['phone'])
        if not phone_valid:
            errors.append(phone_msg)
    
    return errors


def build_preview(batch, mapping, preview_limit=8):
    mapping = normalize_mapping(mapping)
    if not mapping:
        raise ValueError('请先完成 Excel 字段映射。')
    headers, rows = load_excel_rows(batch.file_path)
    preview_rows = []
    valid_count = 0
    invalid_count = 0
    all_errors = []
    for index, row in enumerate(rows, start=1):
        mapped = map_row_data(row, mapping)
        normalized = normalize_resident_payload(mapped)
        errors = validate_normalized_row(normalized)
        if errors:
            invalid_count += 1
            error_detail = {
                'row_number': index + 1,
                'full_name': normalized['full_name'] or '',
                'identity_number': normalized['identity_number'] or '',
                'messages': errors,
                'original_data': row,
            }
            all_errors.append(error_detail)
        else:
            valid_count += 1
        if len(preview_rows) < preview_limit:
            preview_rows.append(
                {
                    'row_number': index + 1,
                    'full_name': normalized['full_name'],
                    'gender': normalized['gender'],
                    'identity_number': normalized['identity_number'],
                    'birth_date': normalized['birth_date'].isoformat() if normalized['birth_date'] else '',
                    'age': calculate_age(normalized['birth_date']),
                    'phone': normalized['phone'],
                    'head_name': normalized['head_name'],
                    'head_identity_number': normalized['head_identity_number'],
                    'relation_to_head': normalized['relation_to_head'],
                    'village_group': normalized['village_group'],
                    'address': normalized['address'],
                    'household_type': normalized['household_type'],
                    'grid_name': normalized['grid_name'],
                    'status': normalized['status'],
                    'errors': errors,
                }
            )
    
    batch.field_mapping = mapping
    batch.valid_rows = valid_count
    batch.invalid_rows = invalid_count
    batch.error_details = all_errors
    batch.status = ResidentImportBatch.STATUS_PREVIEWED
    batch.save(update_fields=['field_mapping', 'valid_rows', 'invalid_rows', 'error_details', 'status', 'updated_at'])
    
    return {
        'headers': headers,
        'preview_rows': preview_rows,
        'total_rows': len(rows),
        'valid_rows': valid_count,
        'invalid_rows': invalid_count,
        'errors': all_errors[:20],
    }


@transaction.atomic
def generate_household_no():
    prefix = timezone.localtime().strftime('HH%Y%m%d')
    existing = Household.objects.filter(
        household_no__startswith=prefix
    ).select_for_update().order_by('-household_no').first()
    if existing:
        last_num = int(existing.household_no[-4:])
        return f'{prefix}{last_num + 1:04d}'
    return f'{prefix}0001'


def update_household_snapshot(household, normalized):
    household.head_name = normalized['head_name'] or household.head_name or ''
    household.head_identity_number = normalized['head_identity_number'] or household.head_identity_number or ''
    household.head_gender = normalized['gender'] if normalized['is_household_head'] else (household.head_gender or '')
    household.village_group = normalized['village_group'] or household.village_group or ''
    household.address = normalized['address'] or household.address or ''
    household.household_type = normalized['household_type'] or household.household_type or ''
    household.account_type = normalized['account_type'] or household.account_type or ''
    household.grid_name = normalized['grid_name'] or household.grid_name or ''
    household.housing_type = normalized['housing_type'] or household.housing_type or ''


def locate_or_create_household(normalized):
    household = None
    if normalized['head_identity_number']:
        household = Household.objects.filter(
            head_identity_number=normalized['head_identity_number']
        ).first()
    if household is None and normalized['household_no']:
        household = Household.objects.filter(household_no=normalized['household_no']).first()
    if household is None and normalized['is_household_head'] and normalized['identity_number']:
        household = Household.objects.filter(
            Q(head_identity_number=normalized['identity_number']) | Q(household_no=normalized['household_no'])
        ).first()

    if household is None:
        household = Household(
            household_no=normalized['household_no'] or generate_household_no(),
        )

    if not household.household_no:
        household.household_no = generate_household_no()
    if not normalized['head_identity_number'] and normalized['is_household_head']:
        normalized['head_identity_number'] = normalized['identity_number']
    if not normalized['head_name'] and normalized['is_household_head']:
        normalized['head_name'] = normalized['full_name']

    update_household_snapshot(household, normalized)
    household.save()
    return household


def save_resident_from_normalized(normalized):
    household = locate_or_create_household(normalized)
    # 确保空字符串正确处理，并且布尔值严格转换
    defaults = {
        'household': household,
        'full_name': normalized['full_name'],
        'gender': normalized['gender'] or '',
        'birth_date': normalized['birth_date'],
        'ethnicity': normalized['ethnicity'] or '',
        'phone': normalized['phone'] or '',
        'relation_to_head': normalized['relation_to_head'] or '',
        'village_group': normalized['village_group'] or '',
        'address': normalized['address'] or '',
        'household_type': normalized['household_type'] or '',
        'account_type': normalized['account_type'] or '',
        'grid_name': normalized['grid_name'] or '',
        'political_status': normalized['political_status'] or '',
        'marital_status': normalized['marital_status'] or '',
        'health_status': normalized['health_status'] or '',
        'residency_status': normalized['residency_status'] or '',
        'education_level': normalized['education_level'] or '',
        'occupation': normalized['occupation'] or '',
        'bank_account': normalized['bank_account'] or '',
        'bank_name': normalized['bank_name'] or '',
        'military_status': normalized['military_status'] or '',
        'status': normalized['status'] or '正常',
        'is_household_head': bool(normalized['is_household_head']),
    }
    resident, created = Resident.objects.update_or_create(
        identity_number=normalized['identity_number'],
        defaults=defaults,
    )

    if normalized['is_household_head']:
        household.head_name = resident.full_name
        household.head_identity_number = resident.identity_number
        household.head_gender = resident.gender
        household.save(update_fields=['head_name', 'head_identity_number', 'head_gender', 'updated_at'])
    return resident, created


@transaction.atomic
def commit_import(batch, mapping):
    build_preview(batch, mapping, preview_limit=20)
    headers, rows = load_excel_rows(batch.file_path)
    created_count = 0
    updated_count = 0
    skipped = []
    error_details = []

    for index, row in enumerate(rows, start=1):
        mapped = map_row_data(row, normalize_mapping(mapping))
        normalized = normalize_resident_payload(mapped)
        errors = validate_normalized_row(normalized)
        if errors:
            error_detail = {
                'row_number': index + 1,
                'full_name': normalized['full_name'] or '',
                'identity_number': normalized['identity_number'] or '',
                'messages': errors,
                'original_data': row,
            }
            skipped.append(error_detail)
            error_details.append(error_detail)
            continue
        try:
            _, created = save_resident_from_normalized(normalized)
            if created:
                created_count += 1
            else:
                updated_count += 1
        except Exception as e:
            error_detail = {
                'row_number': index + 1,
                'full_name': normalized['full_name'] or '',
                'identity_number': normalized['identity_number'] or '',
                'messages': [f'数据库错误: {str(e)}'],
                'original_data': row,
            }
            skipped.append(error_detail)
            error_details.append(error_detail)

    batch.status = ResidentImportBatch.STATUS_IMPORTED if not skipped else ResidentImportBatch.STATUS_FAILED
    batch.imported_rows = created_count + updated_count
    batch.created_rows = created_count
    batch.updated_rows = updated_count
    batch.error_details = error_details
    batch.save(update_fields=['status', 'imported_rows', 'created_rows', 'updated_rows', 'error_details', 'updated_at'])
    return {
        'total_rows': batch.total_rows,
        'valid_rows': batch.valid_rows,
        'invalid_rows': batch.invalid_rows,
        'created_rows': created_count,
        'updated_rows': updated_count,
        'skipped_rows': len(skipped),
        'errors': skipped[:20],
    }


def generate_error_report_workbook(batch):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = '错误详情'
    
    headers = ['行号', '姓名', '身份证号', '错误信息']
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    for row_idx, error in enumerate(batch.error_details, start=2):
        sheet.cell(row=row_idx, column=1, value=error['row_number'])
        sheet.cell(row=row_idx, column=2, value=error['full_name'])
        sheet.cell(row=row_idx, column=3, value=error['identity_number'])
        sheet.cell(row=row_idx, column=4, value='; '.join(error['messages']))
    
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column].width = adjusted_width
    
    return workbook


def get_population_age_structure():
    from .models import LowIncomeRecord, DisabledPerson

    today = timezone.localdate()
    d14 = date(today.year - 14, today.month, today.day)
    d18 = date(today.year - 18, today.month, today.day)
    d59 = date(today.year - 59, today.month, today.day)

    residents = Resident.objects.filter(status='正常')
    agg = residents.aggregate(
        c14=Count('id', filter=Q(birth_date__gte=d14)),
        c15_18=Count('id', filter=Q(birth_date__lt=d14, birth_date__gte=d18)),
        c19_59=Count('id', filter=Q(birth_date__lt=d18, birth_date__gte=d59)),
        c60=Count('id', filter=Q(birth_date__lt=d59)),
    )

    age_groups = {
        '0-14岁': agg['c14'],
        '15-18岁': agg['c15_18'],
        '19-59岁': agg['c19_59'],
        '60岁及以上': agg['c60'],
    }

    total = sum(age_groups.values())
    result = []
    for group, count in age_groups.items():
        percentage = (count / total * 100) if total > 0 else 0
        result.append({
            'name': group,
            'count': count,
            'percentage': round(percentage, 2)
        })

    low_income_count = LowIncomeRecord.objects.filter(status='在享').count()
    disabled_count = DisabledPerson.objects.filter(status='有效').count()

    return {
        'groups': result,
        'total': total,
        'low_income_count': low_income_count,
        'disabled_count': disabled_count
    }


def calculate_age(birth_date):
    if not birth_date:
        return None
    today = timezone.localdate()
    years = today.year - birth_date.year
    if (today.month, today.day) < (birth_date.month, birth_date.day):
        years -= 1
    return years


def shift_years_safe(target_date, years):
    try:
        return target_date.replace(year=target_date.year + years)
    except ValueError:
        return target_date.replace(month=2, day=28, year=target_date.year + years)


def serialize_resident(resident, index=None):
    household = resident.household
    return {
        'id': resident.id,
        'seq': index,
        'full_name': resident.full_name,
        'gender': resident.gender,
        'identity_number': resident.identity_number,
        'birth_date': resident.birth_date.isoformat() if resident.birth_date else '',
        'age': resident.age,
        'ethnicity': resident.ethnicity,
        'phone': resident.phone,
        'head_name': household.head_name if household else '',
        'head_identity_number': household.head_identity_number if household else '',
        'relation_to_head': resident.relation_to_head,
        'village_group': resident.village_group or (household.village_group if household else ''),
        'address': resident.address or (household.address if household else ''),
        'household_type': resident.household_type or (household.household_type if household else ''),
        'grid_name': resident.grid_name or (household.grid_name if household else ''),
        'political_status': resident.political_status,
        'marital_status': resident.marital_status,
        'health_status': resident.health_status,
        'residency_status': resident.residency_status,
        'education_level': resident.education_level,
        'occupation': resident.occupation,
        'bank_account': resident.bank_account,
        'bank_name': resident.bank_name,
        'military_status': resident.military_status,
        'status': resident.status,
        'is_household_head': resident.is_household_head,
        'household_no': household.household_no if household else '',
    }


def serialize_household_summary(item, index=None):
    return {
        'id': item.id,
        'seq': index,
        'head_name': item.head_name,
        'head_gender': item.head_gender,
        'village_group': item.village_group,
        'address': item.address,
        'head_identity_number': item.head_identity_number,
        'member_count': item.member_count,
        'household_type': item.household_type,
        'grid_name': item.grid_name,
        'household_no': item.household_no,
    }


def apply_detail_filters(queryset, params):
    if params.get('full_name'):
        queryset = queryset.filter(full_name__icontains=params['full_name'])
    if params.get('head_name'):
        queryset = queryset.filter(household__head_name__icontains=params['head_name'])
    if params.get('gender'):
        queryset = queryset.filter(gender=params['gender'])
    if params.get('village_group'):
        queryset = queryset.filter(village_group=params['village_group'])
    if params.get('identity_number'):
        queryset = queryset.filter(identity_number__icontains=params['identity_number'])
    if params.get('phone'):
        queryset = queryset.filter(phone__icontains=params['phone'])
    if params.get('political_status'):
        queryset = queryset.filter(political_status=params['political_status'])
    if params.get('marital_status'):
        queryset = queryset.filter(marital_status=params['marital_status'])
    if params.get('health_status'):
        queryset = queryset.filter(health_status=params['health_status'])
    if params.get('residency_status'):
        queryset = queryset.filter(residency_status=params['residency_status'])
    if params.get('household_type'):
        queryset = queryset.filter(household_type=params['household_type'])
    if params.get('grid_name'):
        queryset = queryset.filter(grid_name=params['grid_name'])
    if params.get('status'):
        queryset = queryset.filter(status=params['status'])
    if params.get('address'):
        queryset = queryset.filter(address__icontains=params['address'])

    birth_year_start = params.get('birth_year_start')
    birth_year_end = params.get('birth_year_end')
    if birth_year_start and birth_year_start.isdigit():
        queryset = queryset.filter(birth_date__year__gte=int(birth_year_start))
    if birth_year_end and birth_year_end.isdigit():
        queryset = queryset.filter(birth_date__year__lte=int(birth_year_end))

    age_min = params.get('age_min')
    age_max = params.get('age_max')
    today = timezone.localdate()
    if age_min and age_min.isdigit():
        latest_birth = shift_years_safe(today, -int(age_min))
        queryset = queryset.filter(birth_date__lte=latest_birth)
    if age_max and age_max.isdigit():
        earliest_birth = shift_years_safe(today, -int(age_max) - 1) + timedelta(days=1)
        queryset = queryset.filter(birth_date__gte=earliest_birth)
    return queryset


def apply_household_filters(queryset, params):
    if params.get('full_name'):
        queryset = queryset.filter(
            Q(head_name__icontains=params['full_name']) | Q(residents__full_name__icontains=params['full_name'])
        )
    if params.get('head_name'):
        queryset = queryset.filter(head_name__icontains=params['head_name'])
    if params.get('gender'):
        queryset = queryset.filter(head_gender=params['gender'])
    if params.get('village_group'):
        queryset = queryset.filter(village_group=params['village_group'])
    if params.get('identity_number'):
        queryset = queryset.filter(head_identity_number__icontains=params['identity_number'])
    if params.get('phone'):
        queryset = queryset.filter(residents__phone__icontains=params['phone'])
    if params.get('household_type'):
        queryset = queryset.filter(household_type=params['household_type'])
    if params.get('grid_name'):
        queryset = queryset.filter(grid_name=params['grid_name'])
    if params.get('address'):
        queryset = queryset.filter(address__icontains=params['address'])
    return queryset


def paginate_queryset(queryset, page_number, page_size):
    paginator = Paginator(queryset, page_size)
    page = paginator.get_page(page_number)
    return paginator, page


def list_residents(params):
    page_number = int(params.get('page', 1) or 1)
    page_size = int(params.get('page_size', 10) or 10)
    queryset = Resident.objects.select_related('household').all()
    queryset = apply_detail_filters(queryset, params)
    paginator, page = paginate_queryset(queryset, page_number, page_size)
    start_index = (page.number - 1) * page_size + 1
    items = [
        serialize_resident(resident, index=start_index + offset)
        for offset, resident in enumerate(page.object_list)
    ]
    return {
        'items': items,
        'pagination': {
            'page': page.number,
            'page_size': page_size,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
        },
    }


def list_households(params):
    page_number = int(params.get('page', 1) or 1)
    page_size = int(params.get('page_size', 10) or 10)
    queryset = Household.objects.annotate(member_count=Count('residents', distinct=True)).order_by('household_no')
    queryset = apply_household_filters(queryset, params).distinct()
    paginator, page = paginate_queryset(queryset, page_number, page_size)
    start_index = (page.number - 1) * page_size + 1
    items = [
        serialize_household_summary(household, index=start_index + offset)
        for offset, household in enumerate(page.object_list)
    ]
    return {
        'items': items,
        'pagination': {
            'page': page.number,
            'page_size': page_size,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
        },
    }


def create_resident(payload):
    normalized = normalize_resident_payload(payload)
    errors = validate_normalized_row(normalized)
    if errors:
        raise ValueError('；'.join(errors))
    resident, created = save_resident_from_normalized(normalized)
    return serialize_resident(resident), created


def get_resident_prefill(params):
    full_name = serialize_scalar(params.get('head_name') or params.get('full_name') or params.get('name'))
    identity_number = normalize_identity_number(params.get('identity_number'))
    if not full_name or not identity_number:
        raise ValueError('请先填写户主姓名和身份证号。')

    resident = Resident.objects.select_related('household').filter(identity_number=identity_number).first()
    household = Household.objects.filter(head_identity_number=identity_number).first()
    matched_name = False

    if resident and resident.full_name == full_name:
        matched_name = True
        household = resident.household or household
    elif household and household.head_name == full_name:
        matched_name = True

    if not matched_name:
        return {'found': False, 'item': None}

    head_resident = None
    if resident and resident.full_name == full_name:
        head_resident = resident
    elif household:
        head_resident = Resident.objects.filter(identity_number=household.head_identity_number).first()

    return {
        'found': True,
        'item': {
            'head_name': household.head_name if household else (head_resident.full_name if head_resident else full_name),
            'household_no': household.household_no if household else '',
            'head_gender': household.head_gender if household else (head_resident.gender if head_resident else ''),
            'ethnicity': head_resident.ethnicity if head_resident else '',
            'head_identity_number': household.head_identity_number if household else identity_number,
            'head_phone': head_resident.phone if head_resident else '',
            'account_type': household.account_type if household else (head_resident.account_type if head_resident else ''),
            'household_type': household.household_type if household else (head_resident.household_type if head_resident else ''),
            'grid_name': household.grid_name if household else (head_resident.grid_name if head_resident else ''),
            'housing_type': household.housing_type if household else '',
            'village_group': household.village_group if household else (head_resident.village_group if head_resident else ''),
            'address': household.address if household else (head_resident.address if head_resident else ''),
            'full_name': head_resident.full_name if head_resident else full_name,
            'gender': head_resident.gender if head_resident else '',
            'birth_date': head_resident.birth_date.isoformat() if head_resident and head_resident.birth_date else '',
            'identity_number': head_resident.identity_number if head_resident else identity_number,
            'relation_to_head': head_resident.relation_to_head if head_resident else '户主',
            'marital_status': head_resident.marital_status if head_resident else '',
            'political_status': head_resident.political_status if head_resident else '',
            'military_status': head_resident.military_status if head_resident else '',
            'bank_account': head_resident.bank_account if head_resident else '',
            'bank_name': head_resident.bank_name if head_resident else '',
            'education_level': head_resident.education_level if head_resident else '',
            'occupation': head_resident.occupation if head_resident else '',
            'phone': head_resident.phone if head_resident else '',
            'health_status': head_resident.health_status if head_resident else '',
            'residency_status': head_resident.residency_status if head_resident else '',
            'notes': head_resident.notes if head_resident else '',
        },
    }


def get_resident_options():
    dynamic_options = {}
    for key in ['village_group', 'household_type', 'grid_name']:
        values = (
            Resident.objects.exclude(**{f'{key}__exact': ''})
            .order_by(key)
            .values_list(key, flat=True)
            .distinct()
        )
        options_list = list(values)
        if key == 'village_group':
            vg_names = list(VillageGroup.objects.values_list('name', flat=True))
            options_list = sorted(list(set(options_list) | set(vg_names)))
        dynamic_options[key] = options_list or OPTION_FIELDS.get(key, [])
    payload = {}
    for key, values in OPTION_FIELDS.items():
        payload[key] = dynamic_options.get(key, values)
    return payload

def cascade_update_village_group_name(old_name, new_name):
    if not old_name or old_name == new_name:
        return
    app_config = apps.get_app_config('village_affairs')
    for model in app_config.get_models():
        field_names = [f.name for f in model._meta.get_fields()]
        if 'village_group' in field_names:
            model.objects.filter(village_group=old_name).update(village_group=new_name)


def build_export_workbook(view, params):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if view == 'household':
        sheet.title = '居民按户汇总'
        columns = HOUSEHOLD_EXPORT_COLUMNS
        data = list_households({**params, 'page': 1, 'page_size': 100000})['items']
    else:
        sheet.title = '居民明细'
        columns = DETAIL_EXPORT_COLUMNS
        data = list_residents({**params, 'page': 1, 'page_size': 100000})['items']

    for col_index, (label, _) in enumerate(columns, start=1):
        sheet.cell(row=1, column=col_index, value=label)
    for row_index, item in enumerate(data, start=2):
        for col_index, (_, key) in enumerate(columns, start=1):
            sheet.cell(row=row_index, column=col_index, value=item.get(key, ''))
    return workbook


def get_resident_detail(resident_id):
    resident = Resident.objects.select_related('household').get(id=resident_id)
    household = resident.household
    household_members = []
    if household:
        household_members = [
            serialize_resident(member)
            for member in Resident.objects.filter(household=household).order_by('id')
        ]
    return {
        'resident': serialize_resident(resident),
        'household': {
            'household_no': household.household_no if household else '',
            'head_name': household.head_name if household else '',
            'head_identity_number': household.head_identity_number if household else '',
            'head_gender': household.head_gender if household else '',
            'village_group': household.village_group if household else '',
            'address': household.address if household else '',
            'household_type': household.household_type if household else '',
            'account_type': household.account_type if household else '',
            'grid_name': household.grid_name if household else '',
            'housing_type': household.housing_type if household else '',
        } if household else None,
        'household_members': household_members,
    }


@transaction.atomic
def delete_resident(resident_id):
    resident = Resident.objects.get(id=resident_id)
    household = resident.household
    resident.delete()
    
    if household:
        remaining_members = Resident.objects.filter(household=household).count()
        if remaining_members == 0:
            household.delete()
            return {'message': '居民已删除，该家庭已无其他成员，家庭信息已一并删除'}
        else:
            if household.head_identity_number == resident.identity_number:
                new_head = Resident.objects.filter(household=household).first()
                if new_head:
                    household.head_name = new_head.full_name
                    household.head_identity_number = new_head.identity_number
                    household.head_gender = new_head.gender
                    new_head.is_household_head = True
                    new_head.relation_to_head = '户主'
                    new_head.save(update_fields=['is_household_head', 'relation_to_head'])
                    household.save(update_fields=['head_name', 'head_identity_number', 'head_gender', 'updated_at'])
                    return {'message': '居民已删除，该居民原为主户，已自动指定新户主'}
            return {'message': '居民已删除'}
    
    return {'message': '居民已删除'}


def serialize_migrant_worker(worker, index=None):
    resident = worker.resident
    return {
        'id': worker.id,
        'seq': index,
        'resident_id': resident.id if resident else None,
        'full_name': worker.full_name,
        'gender': worker.gender,
        'identity_number': worker.identity_number,
        'village_group': worker.village_group,
        'phone': worker.phone,
        'household_type': worker.household_type,
        'work_status': worker.work_status,
        'is_employed': worker.is_employed,
        'work_area': worker.work_area,
        'work_address': worker.work_address,
        'work_industry': worker.work_industry,
        'work_type': worker.work_type,
        'employer': worker.employer,
        'start_date': worker.start_date.isoformat() if worker.start_date else '',
        'expected_return_date': worker.expected_return_date.isoformat() if worker.expected_return_date else '',
        'actual_return_date': worker.actual_return_date.isoformat() if worker.actual_return_date else '',
        'is_special_group': worker.is_special_group,
        'month_income': worker.month_income,
        'year_income': worker.year_income,
        'notes': worker.notes,
        'updated_at': worker.updated_at.isoformat() if worker.updated_at else '',
    }


def list_migrant_workers(params):
    page_number = int(params.get('page', 1) or 1)
    page_size = int(params.get('page_size', 10) or 10)
    queryset = MigrantWorker.objects.select_related('resident').all()

    if params.get('full_name'):
        queryset = queryset.filter(full_name__icontains=params['full_name'])
    if params.get('identity_number'):
        queryset = queryset.filter(identity_number__icontains=params['identity_number'])
    if params.get('village_group'):
        queryset = queryset.filter(village_group=params['village_group'])
    if params.get('work_status') and params['work_status'] != '全部':
        queryset = queryset.filter(work_status=params['work_status'])

    paginator, page = paginate_queryset(queryset, page_number, page_size)
    start_index = (page.number - 1) * page_size + 1
    items = [
        serialize_migrant_worker(worker, index=start_index + offset)
        for offset, worker in enumerate(page.object_list)
    ]
    return {
        'items': items,
        'pagination': {
            'page': page.number,
            'page_size': page_size,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
        },
    }


def get_migrant_stats(village_group=None):
    today = timezone.localdate()
    queryset = Resident.objects.filter(status='正常')
    if village_group and village_group != '全部村组':
        queryset = queryset.filter(village_group=village_group)
    total_residents = queryset.count()

    # 计算劳动力总数 (16-60岁) 使用数据库过滤
    d16 = date(today.year - 16, today.month, today.day)
    d60 = date(today.year - 60, today.month, today.day)
    labor_force = queryset.filter(
        birth_date__lt=d16,
        birth_date__gte=d60,
    ).count()

    # 务工统计
    worker_queryset = MigrantWorker.objects.all()
    if village_group and village_group != '全部村组':
        worker_queryset = worker_queryset.filter(village_group=village_group)

    out_working = worker_queryset.filter(work_status='在外务工').count()
    returned = worker_queryset.filter(work_status='已返乡').count()
    special_group = worker_queryset.filter(is_special_group=True).count()

    return {
        'total_residents': total_residents,
        'labor_force': labor_force,
        'out_working': out_working,
        'returned': returned,
        'special_group': special_group,
    }


def get_employment_trend(view_mode='month'):
    today = timezone.localdate()
    if view_mode == 'month':
        months = [f'{i+1}月' for i in range(12)]
        data = []
        for i in range(12):
            month_num = i + 1
            # 简化处理，返回当前统计数据
            data.append({
                'name': months[i],
                '就业人数': MigrantWorker.objects.filter(work_status='在外务工').count(),
                '劳动力就业率': 75,
                '人口就业比': 45,
            })
        return data
    else:
        years = [f'{y}' for y in range(2020, 2027)]
        data = []
        for year in years:
            data.append({
                'name': year,
                '就业人数': MigrantWorker.objects.filter(work_status='在外务工').count(),
                '劳动力就业率': 75,
                '人口就业比': 45,
            })
        return data


def get_mediation_trend():
    """获取当年纠纷处理趋势数据"""
    from django.db.models.functions import ExtractMonth
    from .models import MediationRecord
    today = date.today()
    year = today.year

    # 按月统计纠纷数量
    monthly_stats = MediationRecord.objects.filter(
        application_date__year=year
    ).annotate(
        month=ExtractMonth('application_date')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')

    # 构建12个月的数据
    stats_dict = {item['month']: item['count'] for item in monthly_stats}
    result = []
    for i in range(1, 13):
        result.append({
            'name': f'{i}月',
            'value': stats_dict.get(i, 0)
        })
    return result


@transaction.atomic
def create_migrant_worker(resident_id, data):
    resident = Resident.objects.get(id=resident_id)
    # 检查是否已有务工信息
    if MigrantWorker.objects.filter(resident=resident).exists():
        raise ValueError('该居民已存在务工信息')

    worker = MigrantWorker.objects.create(
        resident=resident,
        full_name=resident.full_name,
        identity_number=resident.identity_number,
        gender=resident.gender,
        village_group=resident.village_group,
        phone=resident.phone,
        household_type=resident.household_type,
        work_status=data.get('work_status', '在外务工'),
        is_employed=data.get('is_employed', True),
        work_area=data.get('work_area', ''),
        work_address=data.get('work_address', ''),
        work_industry=data.get('work_industry', ''),
        work_type=data.get('work_type', ''),
        employer=data.get('employer', ''),
        start_date=parse_date_value(data.get('start_date', '')) if data.get('start_date') else None,
        expected_return_date=parse_date_value(data.get('expected_return_date', '')) if data.get('expected_return_date') else None,
        actual_return_date=parse_date_value(data.get('actual_return_date', '')) if data.get('actual_return_date') else None,
        is_special_group=data.get('is_special_group', False),
        month_income=data.get('month_income', ''),
        year_income=data.get('year_income', ''),
        notes=data.get('notes', ''),
    )
    return serialize_migrant_worker(worker)


@transaction.atomic
def update_migrant_worker(worker_id, data):
    worker = MigrantWorker.objects.get(id=worker_id)
    if 'work_status' in data:
        worker.work_status = data['work_status']
    if 'is_employed' in data:
        worker.is_employed = data['is_employed']
    if 'work_area' in data:
        worker.work_area = data['work_area']
    if 'work_address' in data:
        worker.work_address = data['work_address']
    if 'work_industry' in data:
        worker.work_industry = data['work_industry']
    if 'work_type' in data:
        worker.work_type = data['work_type']
    if 'employer' in data:
        worker.employer = data['employer']
    if 'start_date' in data:
        worker.start_date = parse_date_value(data['start_date']) if data['start_date'] else None
    if 'expected_return_date' in data:
        worker.expected_return_date = parse_date_value(data['expected_return_date']) if data['expected_return_date'] else None
    if 'actual_return_date' in data:
        worker.actual_return_date = parse_date_value(data['actual_return_date']) if data['actual_return_date'] else None
    if 'is_special_group' in data:
        worker.is_special_group = data['is_special_group']
    if 'month_income' in data:
        worker.month_income = data['month_income']
    if 'year_income' in data:
        worker.year_income = data['year_income']
    if 'notes' in data:
        worker.notes = data['notes']
    worker.save()
    return serialize_migrant_worker(worker)


def delete_migrant_worker(worker_id):
    worker = MigrantWorker.objects.get(id=worker_id)
    worker.delete()
    return {'message': '务工信息已删除'}


def get_migrant_worker_detail(worker_id):
    worker = MigrantWorker.objects.select_related('resident').get(id=worker_id)
    resident = worker.resident
    return {
        'worker': serialize_migrant_worker(worker),
        'resident': serialize_resident(resident) if resident else None,
    }


def serialize_risk_check(risk, index=None):
    return {
        'id': risk.id,
        'seq': index,
        'resident_id': risk.resident_id,
        'full_name': risk.full_name,
        'identity_number': risk.identity_number,
        'head_name': risk.head_name,
        'head_identity_number': risk.head_identity_number,
        'household_type': risk.household_type,
        'risk_level': risk.risk_level,
        'warning_content': risk.warning_content,
        'medical_amount': str(risk.medical_amount) if risk.medical_amount is not None else '',
        'warning_time': risk.warning_time.isoformat() if risk.warning_time else '',
        'alert_time': risk.alert_time.isoformat() if risk.alert_time else '',
        'created_at': risk.created_at.isoformat() if risk.created_at else '',
        'updated_at': risk.updated_at.isoformat() if risk.updated_at else '',
        'village_group': risk.resident.village_group if risk.resident else '',
    }


def list_risk_checks(params):
    page_number = int(params.get('page', 1) or 1)
    page_size = int(params.get('page_size', 10) or 10)
    queryset = RiskCheck.objects.select_related('resident').all()

    if params.get('full_name'):
        queryset = queryset.filter(full_name__icontains=params['full_name'])
    if params.get('identity_number'):
        queryset = queryset.filter(identity_number__icontains=params['identity_number'])
    if params.get('risk_level') and params['risk_level'] != '全部':
        queryset = queryset.filter(risk_level=params['risk_level'])
    if params.get('start_date'):
        queryset = queryset.filter(warning_time__gte=parse_date_value(params['start_date']))
    if params.get('end_date'):
        queryset = queryset.filter(warning_time__lte=parse_date_value(params['end_date']))

    paginator, page = paginate_queryset(queryset, page_number, page_size)
    start_index = (page.number - 1) * page_size + 1
    items = [
        serialize_risk_check(item, index=start_index + offset)
        for offset, item in enumerate(page.object_list)
    ]
    return {
        'items': items,
        'pagination': {
            'page': page.number,
            'page_size': page_size,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
        },
    }


@transaction.atomic
def create_risk_check(data):
    resident_id = data.get('resident_id')
    if not resident_id:
        raise ValueError('必须关联一个居民')
    resident = Resident.objects.select_related('household').get(id=resident_id)
    household = resident.household
    
    risk = RiskCheck.objects.create(
        resident=resident,
        full_name=resident.full_name,
        identity_number=resident.identity_number,
        head_name=household.head_name if household else '',
        head_identity_number=household.head_identity_number if household else '',
        household_type=data.get('household_type', resident.household_type),
        risk_level=data.get('risk_level', ''),
        warning_content=data.get('warning_content', ''),
        medical_amount=data.get('medical_amount') or None,
        warning_time=parse_date_value(data.get('warning_time')) or timezone.localdate(),
    )
    return serialize_risk_check(risk)


def delete_risk_check(risk_id):
    risk = RiskCheck.objects.get(id=risk_id)
    risk.delete()
    return {'message': '风险预警记录已删除'}
