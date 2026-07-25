import json
from datetime import timedelta
from email.header import decode_header
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

import openpyxl
from django.db import ProgrammingError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, TransactionTestCase
from django.utils import timezone

from .models import (
    CareObject,
    DatabaseBackupRecord,
    DisabledPerson,
    FarmlandRecord,
    Household,
    LowIncomeRecord,
    MediationRecord,
    MigrantWorker,
    OrganizationMember,
    OperationLog,
    PartyFeeRecord,
    PartyMember,
    PublicJobRecord,
    ProjectRecord,
    ReminderRule,
    Resident,
    RiskCheck,
    SubsidyRecord,
    TodoReminder,
    UISettings,
)


class ResidentFeatureTests(TestCase):
    maxDiff = None

    def build_excel_file(self, headers, rows, filename='residents.xlsx'):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return SimpleUploadedFile(
            filename,
            stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def upload_import_file(self, headers, rows):
        file = self.build_excel_file(headers, rows)
        response = self.client.post('/api/residents/import/upload/', {'file': file})
        self.assertEqual(response.status_code, 200)
        return response.json()

    def seed_imported_family(self):
        upload_payload = self.upload_import_file(
            headers=['姓名列', '证件列', '性别列', '户主证件列', '关系列', '村组列', '地址列', '户属性列', '网格列', '联系电话列'],
            rows=[
                ['张三', '110101199001010011', '男', '110101199001010011', '户主', '一组', '幸福路1号', '普通户', '网格一', '13800000001'],
                ['李四', '110101201001010022', '女', '110101199001010011', '子女', '一组', '幸福路1号', '普通户', '网格一', '13800000002'],
            ],
        )
        mapping = {
            'full_name': '姓名列',
            'identity_number': '证件列',
            'gender': '性别列',
            'head_identity_number': '户主证件列',
            'relation_to_head': '关系列',
            'village_group': '村组列',
            'address': '地址列',
            'household_type': '户属性列',
            'grid_name': '网格列',
            'phone': '联系电话列',
        }
        return upload_payload, mapping

    def test_import_rejects_non_excel_file(self):
        bad_file = SimpleUploadedFile('invalid.txt', b'not-excel', content_type='text/plain')
        response = self.client.post('/api/residents/import/upload/', {'file': bad_file})
        self.assertEqual(response.status_code, 400)
        self.assertIn('仅支持上传', response.json()['message'])

    def test_population_age_structure_and_stats(self):
        Resident.objects.create(
            full_name='统计居民1',
            identity_number='110101199001010011',
            birth_date='1990-01-01',
            status='正常',
        )
        resident2 = Resident.objects.create(
            full_name='统计居民2',
            identity_number='110101199001010022',
            birth_date='1990-01-01',
            status='正常',
        )
        DisabledPerson.objects.create(
            resident=resident2,
            full_name='统计居民2',
            identity_number='110101199001010022',
            status='有效'
        )
        LowIncomeRecord.objects.create(
            resident=resident2,
            full_name='统计居民2',
            identity_number='110101199001010022',
            status='在享'
        )
        
        response = self.client.get('/api/population/age-structure/')
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['low_income_count'], 1)
        self.assertEqual(data['disabled_count'], 1)

    def test_ui_settings_get_and_update(self):
        get_response = self.client.get('/api/settings/ui/')
        self.assertEqual(get_response.status_code, 200)
        get_payload = get_response.json()
        self.assertEqual(get_payload['systemTitle'], '村务管理系统')
        self.assertEqual(get_payload['logoMode'], 'both')
        self.assertEqual(UISettings.objects.count(), 1)

        put_response = self.client.put(
            '/api/settings/ui/',
            data=json.dumps(
                {
                    'systemTitle': '智慧村务平台',
                    'logoMode': 'text',
                    'logoText': '智慧村务',
                    'logoImage': 'data:image/png;base64,logo',
                    'favicon': 'data:image/png;base64,favicon',
                    'villageOverview': '这里是新的村情概况',
                    'villageImage': 'data:image/png;base64,village',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(put_response.status_code, 200)
        put_payload = put_response.json()['item']
        self.assertEqual(put_payload['systemTitle'], '智慧村务平台')
        self.assertEqual(put_payload['logoMode'], 'text')
        self.assertEqual(put_payload['villageOverview'], '这里是新的村情概况')

        setting = UISettings.objects.get()
        self.assertEqual(setting.system_title, '智慧村务平台')
        self.assertEqual(setting.logo_mode, 'text')
        self.assertEqual(setting.logo_text, '智慧村务')

    def test_operation_log_list_and_cleanup(self):
        self.client.put(
            '/api/settings/ui/',
            data=json.dumps(
                {
                    'systemTitle': '日志测试平台',
                    'logoMode': 'text',
                    'logoText': '日志测试',
                    'villageOverview': '日志测试概况',
                }
            ),
            content_type='application/json',
        )

        old_log = OperationLog.objects.create(
            module='系统设置',
            action='修改',
            operator='系统管理员',
            summary='过期日志',
            result='成功',
        )
        OperationLog.objects.filter(id=old_log.id).update(created_at=timezone.now() - timedelta(days=120))

        list_response = self.client.get('/api/operation-logs/?module=系统设置&result=成功')
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertGreaterEqual(list_payload['pagination']['total'], 1)
        self.assertIn('系统设置', list_payload['filter_options']['modules'])

        cleanup_response = self.client.post(
            '/api/operation-logs/cleanup/',
            data=json.dumps({'retention_days': 90}),
            content_type='application/json',
        )
        self.assertEqual(cleanup_response.status_code, 200)
        cleanup_payload = cleanup_response.json()
        self.assertGreaterEqual(cleanup_payload['deleted_count'], 1)
        self.assertFalse(OperationLog.objects.filter(id=old_log.id).exists())

    def test_database_backup_create_list_download_and_delete(self):
        with TemporaryDirectory() as temp_dir, patch(
            'village_affairs.services.get_database_backup_dir',
            return_value=Path(temp_dir),
        ):
            create_response = self.client.post(
                '/api/data-security/backups/create/',
                data=json.dumps({}),
                content_type='application/json',
            )
            self.assertEqual(create_response.status_code, 201)
            create_payload = create_response.json()
            backup_id = create_payload['item']['id']
            self.assertEqual(create_payload['item']['status'], '成功')
            self.assertTrue(Path(create_payload['item']['file_path']).exists())

            list_response = self.client.get('/api/data-security/backups/')
            self.assertEqual(list_response.status_code, 200)
            list_payload = list_response.json()
            self.assertEqual(list_payload['summary']['total'], 1)
            self.assertEqual(list_payload['items'][0]['id'], backup_id)
            self.assertTrue(list_payload['backup_dir'].endswith(temp_dir))

            download_response = self.client.get(f'/api/data-security/backups/{backup_id}/download/')
            self.assertEqual(download_response.status_code, 200)
            self.assertIn('application/json', download_response['Content-Type'])
            self.assertIn('attachment; filename=', download_response['Content-Disposition'])

            delete_response = self.client.delete(f'/api/data-security/backups/{backup_id}/')
            self.assertEqual(delete_response.status_code, 200)
            self.assertFalse(DatabaseBackupRecord.objects.filter(id=backup_id).exists())

    def test_excel_mapping_preview_commit_and_household_grouping(self):
        upload_payload, mapping = self.seed_imported_family()

        preview_response = self.client.post(
            '/api/residents/import/preview/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(preview_response.status_code, 200)
        preview_payload = preview_response.json()
        self.assertEqual(preview_payload['total_rows'], 2)
        self.assertEqual(preview_payload['valid_rows'], 2)
        self.assertEqual(preview_payload['invalid_rows'], 0)
        self.assertEqual(preview_payload['preview_rows'][0]['full_name'], '张三')
        self.assertEqual(preview_payload['preview_rows'][1]['head_identity_number'], '110101199001010011')

        commit_response = self.client.post(
            '/api/residents/import/commit/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(commit_response.status_code, 200)
        commit_payload = commit_response.json()
        self.assertEqual(commit_payload['created_rows'], 2)
        self.assertEqual(commit_payload['updated_rows'], 0)
        self.assertEqual(commit_payload['skipped_rows'], 0)

        self.assertEqual(Household.objects.count(), 1)
        self.assertEqual(Resident.objects.count(), 2)

        household = Household.objects.get()
        head = Resident.objects.get(identity_number='110101199001010011')
        member = Resident.objects.get(identity_number='110101201001010022')
        self.assertEqual(household.head_identity_number, '110101199001010011')
        self.assertEqual(household.head_name, '张三')
        self.assertEqual(head.household_id, household.id)
        self.assertEqual(member.household_id, household.id)
        self.assertTrue(head.is_household_head)
        self.assertFalse(member.is_household_head)

    def test_household_summary_and_export(self):
        upload_payload, mapping = self.seed_imported_family()
        self.client.post(
            '/api/residents/import/commit/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )

        summary_response = self.client.get('/api/residents/households/')
        self.assertEqual(summary_response.status_code, 200)
        summary_payload = summary_response.json()
        self.assertEqual(summary_payload['pagination']['total'], 1)
        self.assertEqual(summary_payload['items'][0]['head_name'], '张三')
        self.assertEqual(summary_payload['items'][0]['member_count'], 2)

        export_response = self.client.get('/api/residents/export/?view=household')
        self.assertEqual(export_response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            export_response['Content-Type'],
        )

    def test_create_resident_and_query(self):
        response = self.client.post(
            '/api/residents/',
            data=json.dumps(
                {
                    'head_name': '王五',
                    'head_identity_number': '110101198505050033',
                    'full_name': '王五',
                    'gender': '男',
                    'identity_number': '110101198505050033',
                    'relation_to_head': '户主',
                    'village_group': '二组',
                    'address': '团结路8号',
                    'household_type': '普通户',
                    'grid_name': '网格二',
                    'status': '正常',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(Resident.objects.count(), 1)
        self.assertEqual(Household.objects.count(), 1)

        query_response = self.client.get('/api/residents/?full_name=王五')
        self.assertEqual(query_response.status_code, 200)
        payload = query_response.json()
        self.assertEqual(payload['pagination']['total'], 1)
        self.assertEqual(payload['items'][0]['full_name'], '王五')

    def test_import_commit_handles_non_head_rows_without_boolean_errors(self):
        upload_payload = self.upload_import_file(
            headers=['姓名', '身份证号', '户主身份证号', '与户主关系'],
            rows=[
                ['赵六', '110101199202020044', '110101199202020044', '户主'],
                ['赵小六', '110101202002020055', '110101199202020044', ''],
            ],
        )
        mapping = {
            'full_name': '姓名',
            'identity_number': '身份证号',
            'head_identity_number': '户主身份证号',
            'relation_to_head': '与户主关系',
        }

        response = self.client.post(
            '/api/residents/import/commit/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['created_rows'], 2)
        self.assertEqual(payload['updated_rows'], 0)
        self.assertEqual(payload['skipped_rows'], 0)

        member = Resident.objects.get(identity_number='110101202002020055')
        self.assertFalse(member.is_household_head)

    def test_bulk_delete_endpoints_for_migrant_worker_and_risk_check(self):
        resident = Resident.objects.create(
            full_name='周七',
            gender='男',
            identity_number='110101198808080066',
            relation_to_head='户主',
            status='正常',
        )
        worker = MigrantWorker.objects.create(
            resident=resident,
            full_name=resident.full_name,
            identity_number=resident.identity_number,
            gender=resident.gender,
            work_status='在外务工',
        )
        risk = RiskCheck.objects.create(
            resident=resident,
            full_name=resident.full_name,
            identity_number=resident.identity_number,
            risk_level='中风险',
            warning_time='2026-07-03',
        )

        worker_response = self.client.post(
            '/api/migrant-workers/bulk-delete/',
            data=json.dumps({'ids': [worker.id]}),
            content_type='application/json',
        )
        risk_response = self.client.post(
            '/api/risk-checks/bulk-delete/',
            data=json.dumps({'ids': [risk.id]}),
            content_type='application/json',
        )

        self.assertEqual(worker_response.status_code, 200)
        self.assertEqual(risk_response.status_code, 200)
        self.assertFalse(MigrantWorker.objects.filter(id=worker.id).exists())
        self.assertFalse(RiskCheck.objects.filter(id=risk.id).exists())

    def test_risk_check_import_supports_field_mapping_preview_and_commit(self):
        resident = Resident.objects.create(
            full_name='风险对象',
            gender='女',
            identity_number='110101199505050077',
            relation_to_head='户主',
            village_group='三组',
            household_type='低保户',
            status='正常',
        )
        file = self.build_excel_file(
            headers=['证件列', '姓名列', '等级列', '时间列', '内容列', '金额列'],
            rows=[['110101199505050077', '风险对象', '高风险', '2026-07-04', '医疗支出过高', '1234.5']],
            filename='risk-check.xlsx',
        )

        upload_response = self.client.post('/api/risk-checks/import/upload/', {'file': file})
        self.assertEqual(upload_response.status_code, 200)
        upload_payload = upload_response.json()
        mapping = {
            'identity_number': '证件列',
            'full_name': '姓名列',
            'risk_level': '等级列',
            'warning_time': '时间列',
            'warning_content': '内容列',
            'medical_amount': '金额列',
        }

        preview_response = self.client.post(
            '/api/risk-checks/import/preview/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(preview_response.status_code, 200)
        preview_payload = preview_response.json()
        self.assertEqual(preview_payload['valid_rows'], 1)
        self.assertEqual(preview_payload['invalid_rows'], 0)
        self.assertEqual(preview_payload['preview_rows'][0]['risk_level'], '高风险')

        commit_response = self.client.post(
            '/api/risk-checks/import/commit/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(commit_response.status_code, 200)
        commit_payload = commit_response.json()
        self.assertEqual(commit_payload['created_rows'], 1)
        self.assertEqual(commit_payload['updated_rows'], 0)
        self.assertEqual(commit_payload['skipped_rows'], 0)

        risk = RiskCheck.objects.get(identity_number='110101199505050077')
        self.assertEqual(risk.resident_id, resident.id)
        self.assertEqual(str(risk.medical_amount), '1234.50')
        self.assertEqual(risk.warning_content, '医疗支出过高')

    def test_low_income_create_list_and_household_summary(self):
        household = Household.objects.create(
            household_no='H001',
            head_name='低保户主',
            head_identity_number='110101198001010088',
            village_group='二组',
            household_type='低保户',
        )
        resident = Resident.objects.create(
            household=household,
            full_name='低保对象',
            gender='男',
            identity_number='110101199909090099',
            relation_to_head='子女',
            village_group='二组',
            household_type='低保户',
            phone='13800009999',
            status='正常',
        )

        create_response = self.client.post(
            '/api/low-income/create/',
            data=json.dumps(
                {
                    'resident_id': resident.id,
                    'policy_type': '低保',
                    'benefit_level': 'A档',
                    'subsidy_amount': '380',
                    'subsidy_cycle': '按月',
                    'start_date': '2026-07-01',
                    'household_member_count': 3,
                    'beneficiary_count': 1,
                    'household_month_amount': '380',
                    'status': '在享',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 201)

        detail_response = self.client.get('/api/low-income/?full_name=低保对象')
        self.assertEqual(detail_response.status_code, 200)
        detail_payload = detail_response.json()
        self.assertEqual(detail_payload['pagination']['total'], 1)
        self.assertEqual(detail_payload['items'][0]['policy_type'], '低保')
        self.assertEqual(detail_payload['items'][0]['head_name'], '低保户主')

        household_response = self.client.get('/api/low-income/households/?status=在享')
        self.assertEqual(household_response.status_code, 200)
        household_payload = household_response.json()
        self.assertEqual(household_payload['pagination']['total'], 1)
        self.assertEqual(household_payload['items'][0]['beneficiary_count'], 1)
        self.assertEqual(household_payload['items'][0]['policy_type'], '低保')

    def test_disabled_create_and_list(self):
        resident = Resident.objects.create(
            full_name='残疾测试',
            identity_number='110101199901010011',
            village_group='一组',
            phone='13800001111',
            status='正常',
        )
        create_response = self.client.post(
            '/api/disabled/create/',
            data=json.dumps(
                {
                    'resident_id': resident.id,
                    'disability_type': '肢体残疾',
                    'disability_level': '一级',
                    'disability_card_number': '11010119990101001144',
                    'status': '有效',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 201)

        list_response = self.client.get('/api/disabled/?full_name=残疾测试')
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(list_payload['pagination']['total'], 1)
        self.assertEqual(list_payload['items'][0]['disability_type'], '肢体残疾')
        self.assertEqual(list_payload['items'][0]['disability_level'], '一级')

    def test_subsidy_create_and_list(self):
        resident = Resident.objects.create(
            full_name='补贴测试',
            identity_number='110101199902020022',
            village_group='一组',
            bank_account='6217000012345678901',
            status='正常',
        )
        create_response = self.client.post(
            '/api/subsidies/create/',
            data=json.dumps(
                {
                    'resident_id': resident.id,
                    'grant_year': 2026,
                    'subsidy_type': '耕地地力保护补贴',
                    'subsidy_item': '耕地补贴',
                    'declared_amount': '100.00',
                    'payment_status': '待发放',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 201)

        list_response = self.client.get(
            f'/api/subsidies/?full_name=补贴测试&identity_number={resident.identity_number}&payment_status=待发放&village_group=一组'
        )
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(list_payload['pagination']['total'], 1)
        self.assertEqual(list_payload['items'][0]['subsidy_type'], '耕地地力保护补贴')
        self.assertEqual(list_payload['items'][0]['declared_amount'], '100.00')
        self.assertEqual(list_payload['items'][0]['bank_account'], '6217000012345678901')

    def test_subsidy_import_supports_field_mapping_preview_and_commit(self):
        resident = Resident.objects.create(
            full_name='导入补贴对象',
            identity_number='110101199904040044',
            village_group='三组',
            bank_account='6217000099998888777',
            status='正常',
        )
        file = self.build_excel_file(
            headers=['年度列', '补贴类型列', '身份证列', '金额列', '状态列'],
            rows=[['2026', '雨露计划补助', resident.identity_number, '300.50', '已发放']],
            filename='subsidies.xlsx',
        )

        upload_response = self.client.post('/api/subsidies/import/upload/', {'file': file})
        self.assertEqual(upload_response.status_code, 200)
        upload_payload = upload_response.json()
        mapping = {
            'grant_year': '年度列',
            'subsidy_type': '补贴类型列',
            'identity_number': '身份证列',
            'declared_amount': '金额列',
            'payment_status': '状态列',
        }

        preview_response = self.client.post(
            '/api/subsidies/import/preview/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(preview_response.status_code, 200)
        preview_payload = preview_response.json()
        self.assertEqual(preview_payload['valid_rows'], 1)
        self.assertEqual(preview_payload['invalid_rows'], 0)
        self.assertEqual(preview_payload['preview_rows'][0]['subsidy_type'], '雨露计划补助')

        commit_response = self.client.post(
            '/api/subsidies/import/commit/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(commit_response.status_code, 200)
        commit_payload = commit_response.json()
        self.assertEqual(commit_payload['created_rows'], 1)
        self.assertEqual(commit_payload['updated_rows'], 0)

        item = SubsidyRecord.objects.get(identity_number=resident.identity_number, subsidy_type='雨露计划补助')
        self.assertEqual(item.full_name, '导入补贴对象')
        self.assertEqual(item.village_group, '三组')
        self.assertEqual(item.bank_account, '6217000099998888777')
        self.assertEqual(str(item.declared_amount), '300.50')

    def test_public_job_create_and_list(self):
        resident = Resident.objects.create(
            full_name='岗位测试',
            identity_number='110101199905050055',
            village_group='二组',
            phone='13800002222',
            household_type='脱贫户',
            status='正常',
        )
        LowIncomeRecord.objects.create(
            resident=resident,
            household=resident.household,
            full_name=resident.full_name,
            identity_number=resident.identity_number,
            gender=resident.gender,
            ethnicity=resident.ethnicity,
            phone=resident.phone,
            head_name=resident.full_name,
            relation_to_head=resident.relation_to_head,
            village_group=resident.village_group,
            policy_type='低保',
            benefit_level='A档',
            household_member_count=1,
            beneficiary_count=1,
            status='在享',
        )

        create_response = self.client.post(
            '/api/public-jobs/create/',
            data=json.dumps(
                {
                    'resident_id': resident.id,
                    'job_name': '保洁岗',
                    'department': '村委会',
                    'start_date': '2026-07-01',
                    'end_date': '2026-12-31',
                    'subsidy_amount': '500',
                    'required_attendance_days': 22,
                    'actual_attendance_days': 20,
                    'status': '在岗',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 201)

        list_response = self.client.get(f'/api/public-jobs/?full_name=岗位测试&identity_number={resident.identity_number}&status=在岗')
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(list_payload['pagination']['total'], 1)
        self.assertEqual(list_payload['items'][0]['job_name'], '保洁岗')
        self.assertEqual(list_payload['items'][0]['department'], '村委会')
        self.assertEqual(list_payload['items'][0]['low_income_type'], '低保')

    def test_public_job_import_supports_field_mapping_preview_and_commit(self):
        resident = Resident.objects.create(
            full_name='导入岗位对象',
            identity_number='110101199906060066',
            village_group='五组',
            phone='13800003333',
            status='正常',
        )
        file = self.build_excel_file(
            headers=['身份证列', '岗位列', '部门列', '开始列', '状态列', '补贴列'],
            rows=[[resident.identity_number, '护林岗', '人社部门', '2026-07-01', '在岗', '650']],
            filename='public_jobs.xlsx',
        )

        upload_response = self.client.post('/api/public-jobs/import/upload/', {'file': file})
        self.assertEqual(upload_response.status_code, 200)
        upload_payload = upload_response.json()
        mapping = {
            'identity_number': '身份证列',
            'job_name': '岗位列',
            'department': '部门列',
            'start_date': '开始列',
            'status': '状态列',
            'subsidy_amount': '补贴列',
        }

        preview_response = self.client.post(
            '/api/public-jobs/import/preview/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(preview_response.status_code, 200)
        preview_payload = preview_response.json()
        self.assertEqual(preview_payload['valid_rows'], 1)
        self.assertEqual(preview_payload['preview_rows'][0]['job_name'], '护林岗')

        commit_response = self.client.post(
            '/api/public-jobs/import/commit/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(commit_response.status_code, 200)
        commit_payload = commit_response.json()
        self.assertEqual(commit_payload['created_rows'], 1)

        item = PublicJobRecord.objects.get(identity_number=resident.identity_number, job_name='护林岗')
        self.assertEqual(item.full_name, '导入岗位对象')
        self.assertEqual(item.department, '人社部门')
        self.assertEqual(str(item.subsidy_amount), '650.00')

    def test_care_object_create_detail_update_and_list(self):
        resident = Resident.objects.create(
            full_name='关爱测试',
            gender='女',
            ethnicity='苗族',
            identity_number='110101199907070077',
            village_group='三组',
            address='温暖路7号',
            phone='13800004444',
            status='正常',
        )

        create_response = self.client.post(
            '/api/care-objects/create/',
            data=json.dumps(
                {
                    'resident_id': resident.id,
                    'care_type': '留守儿童',
                    'care_level': '重点',
                    'caregiver_name': '李干部',
                    'caregiver_phone': '13900001111',
                    'notes': '每周走访',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 201)
        record_id = create_response.json()['item']['id']

        list_response = self.client.get(
            f'/api/care-objects/?full_name=关爱测试&identity_number={resident.identity_number}&care_type=留守儿童&village_group=三组'
        )
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(list_payload['pagination']['total'], 1)
        self.assertEqual(list_payload['items'][0]['care_level'], '重点')
        self.assertEqual(list_payload['items'][0]['caregiver_name'], '李干部')

        detail_response = self.client.get(f'/api/care-objects/{record_id}/')
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(detail_response.json()['address'], '温暖路7号')

        update_response = self.client.put(
            f'/api/care-objects/{record_id}/',
            data=json.dumps(
                {
                    'care_type': '孤寡老人',
                    'care_level': '一般',
                    'caregiver_name': '王网格员',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.json()['item']['care_type'], '孤寡老人')
        self.assertEqual(update_response.json()['item']['caregiver_name'], '王网格员')

        record = CareObject.objects.get(id=record_id)
        self.assertEqual(record.full_name, '关爱测试')
        self.assertEqual(record.care_level, '一般')

    def test_care_object_import_supports_field_mapping_preview_commit_and_bulk_delete(self):
        resident = Resident.objects.create(
            full_name='导入关爱对象',
            gender='男',
            identity_number='110101199908080088',
            village_group='四组',
            phone='13800005555',
            status='正常',
        )
        file = self.build_excel_file(
            headers=['身份证列', '类型列', '等级列', '人员列', '电话列'],
            rows=[[resident.identity_number, '留守妇女', '重点', '张干部', '13900002222']],
            filename='care_objects.xlsx',
        )

        upload_response = self.client.post('/api/care-objects/import/upload/', {'file': file})
        self.assertEqual(upload_response.status_code, 200)
        upload_payload = upload_response.json()
        mapping = {
            'identity_number': '身份证列',
            'care_type': '类型列',
            'care_level': '等级列',
            'caregiver_name': '人员列',
            'caregiver_phone': '电话列',
        }

        preview_response = self.client.post(
            '/api/care-objects/import/preview/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(preview_response.status_code, 200)
        preview_payload = preview_response.json()
        self.assertEqual(preview_payload['valid_rows'], 1)
        self.assertEqual(preview_payload['preview_rows'][0]['care_type'], '留守妇女')

        commit_response = self.client.post(
            '/api/care-objects/import/commit/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(commit_response.status_code, 200)
        commit_payload = commit_response.json()
        self.assertEqual(commit_payload['created_rows'], 1)

        record = CareObject.objects.get(identity_number=resident.identity_number)
        self.assertEqual(record.full_name, '导入关爱对象')
        self.assertEqual(record.caregiver_name, '张干部')

        bulk_delete_response = self.client.post(
            '/api/care-objects/bulk-delete/',
            data=json.dumps({'ids': [record.id]}),
            content_type='application/json',
        )
        self.assertEqual(bulk_delete_response.status_code, 200)
        self.assertFalse(CareObject.objects.filter(id=record.id).exists())

    def test_project_create_query_update_delete_and_export(self):
        create_response = self.client.post(
            '/api/projects/create/',
            data=json.dumps(
                {
                    'project_name': '幸福路硬化提升',
                    'project_source': '乡村振兴项目库',
                    'project_type': '基础设施',
                    'secondary_type': '道路建设',
                    'project_status': '规划中',
                    'planning_year': 2026,
                    'implementation_year': 2027,
                    'included_in_plan': True,
                    'total_investment': '320.50',
                    'settled_amount': '0',
                    'audited_amount': '0',
                    'responsible_person': '张主任',
                    'project_location': '一组幸福路',
                    'project_description': '道路硬化及排水沟整治',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 201)
        project_id = create_response.json()['item']['id']

        list_response = self.client.get('/api/projects/?keyword=幸福路&project_status=规划中&project_type=基础设施')
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(list_payload['pagination']['total'], 1)
        self.assertEqual(list_payload['items'][0]['project_name'], '幸福路硬化提升')
        self.assertEqual(list_payload['items'][0]['included_in_plan_label'], '是')
        self.assertEqual(list_payload['stats']['total_investment'], '320.50')

        detail_response = self.client.get(f'/api/projects/{project_id}/')
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(detail_response.json()['project_location'], '一组幸福路')

        update_response = self.client.put(
            f'/api/projects/{project_id}/',
            data=json.dumps(
                {
                    'project_status': '实施中',
                    'settled_amount': '120.00',
                    'responsible_person': '李书记',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.json()['item']['project_status'], '实施中')
        self.assertEqual(update_response.json()['item']['settled_amount'], '120.00')

        export_response = self.client.get('/api/projects/export/?project_status=实施中')
        self.assertEqual(export_response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            export_response['Content-Type'],
        )

        delete_response = self.client.delete(f'/api/projects/{project_id}/')
        self.assertEqual(delete_response.status_code, 200)
        self.assertFalse(ProjectRecord.objects.filter(id=project_id).exists())

    def test_project_list_returns_actionable_message_when_project_tables_are_missing(self):
        with patch(
            'village_affairs.views.list_projects',
            side_effect=ProgrammingError(
                '(1146, "Table \'country_manage_system.village_affairs_projectrecord\' doesn\'t exist")'
            ),
        ):
            response = self.client.get('/api/projects/')

        self.assertEqual(response.status_code, 503)
        payload = response.json()
        self.assertEqual(payload['error_code'], 'project_tables_missing')
        self.assertIn('manage.py migrate', payload['message'])

    def test_project_import_supports_field_mapping_preview_and_commit(self):
        file = self.build_excel_file(
            headers=['项目名称列', '来源列', '类型列', '状态列', '规划列', '实施列', '计划列', '投资列', '地点列'],
            rows=[['高标准农田提升', '财政衔接资金项目库', '基础设施', '已完成', '2025', '2026', '是', '456.78', '三组北片区']],
            filename='projects.xlsx',
        )

        upload_response = self.client.post('/api/projects/import/upload/', {'file': file})
        self.assertEqual(upload_response.status_code, 200)
        upload_payload = upload_response.json()
        mapping = {
            'project_name': '项目名称列',
            'project_source': '来源列',
            'project_type': '类型列',
            'project_status': '状态列',
            'planning_year': '规划列',
            'implementation_year': '实施列',
            'included_in_plan': '计划列',
            'total_investment': '投资列',
            'project_location': '地点列',
        }

        preview_response = self.client.post(
            '/api/projects/import/preview/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(preview_response.status_code, 200)
        preview_payload = preview_response.json()
        self.assertEqual(preview_payload['valid_rows'], 1)
        self.assertEqual(preview_payload['invalid_rows'], 0)
        self.assertEqual(preview_payload['preview_rows'][0]['project_name'], '高标准农田提升')

        commit_response = self.client.post(
            '/api/projects/import/commit/',
            data=json.dumps({'batch_id': upload_payload['batch_id'], 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(commit_response.status_code, 200)
        commit_payload = commit_response.json()
        self.assertEqual(commit_payload['created_rows'], 1)
        self.assertEqual(commit_payload['updated_rows'], 0)

        project = ProjectRecord.objects.get(project_name='高标准农田提升')
        self.assertEqual(project.project_source, '财政衔接资金项目库')
        self.assertTrue(project.included_in_plan)
        self.assertEqual(str(project.total_investment), '456.78')
        self.assertEqual(project.project_location, '三组北片区')

    def test_mediation_create_query_update_export_document_and_delete(self):
        create_response = self.client.post(
            '/api/mediations/create/',
            data=json.dumps(
                {
                    'dispute_type': '婚姻家庭纠纷',
                    'status': '进行中',
                    'application_date': '2026-07-07',
                    'occurrence_date': '2026-07-05',
                    'occurrence_location': '一组文化广场',
                    'applicants': [
                        {
                            'name': '张三',
                            'gender': '男',
                            'ethnicity': '汉族',
                            'identity_number': '110101199001010011',
                            'phone': '13800000001',
                            'occupation': '务工',
                            'address': '幸福路1号',
                        }
                    ],
                    'respondents': [
                        {
                            'name': '李四',
                            'gender': '女',
                            'ethnicity': '汉族',
                            'identity_number': '110101199202020022',
                            'phone': '13800000002',
                            'occupation': '个体',
                            'address': '幸福路2号',
                        }
                    ],
                    'dispute_summary': '因宅基地边界问题引发邻里纠纷。',
                    'application_requests': ['厘清边界范围', '恢复原状', '协商补偿'],
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 201)
        created_item = create_response.json()['item']
        record_id = created_item['id']
        self.assertEqual(created_item['archive_number'], 'TJ-001')
        self.assertEqual(created_item['applicant_names'], '张三')
        self.assertEqual(created_item['respondent_names'], '李四')

        next_number_response = self.client.get('/api/mediations/next-archive-number/')
        self.assertEqual(next_number_response.status_code, 200)
        self.assertEqual(next_number_response.json()['archive_number'], 'TJ-002')

        list_response = self.client.get('/api/mediations/?archive_number=TJ-001&dispute_type=婚姻家庭纠纷&status=进行中')
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(list_payload['pagination']['total'], 1)
        self.assertEqual(list_payload['items'][0]['occurrence_location'], '一组文化广场')

        detail_response = self.client.get(f'/api/mediations/{record_id}/')
        self.assertEqual(detail_response.status_code, 200)
        detail_payload = detail_response.json()
        self.assertEqual(detail_payload['application_requests'][0], '厘清边界范围')
        self.assertEqual(detail_payload['respondents'][0]['name'], '李四')

        update_response = self.client.put(
            f'/api/mediations/{record_id}/',
            data=json.dumps(
                {
                    'archive_number': 'TJ-001',
                    'dispute_type': '婚姻家庭纠纷',
                    'status': '已调解',
                    'application_date': '2026-07-07',
                    'occurrence_date': '2026-07-05',
                    'occurrence_location': '村委会调解室',
                    'applicants': detail_payload['applicants'],
                    'respondents': detail_payload['respondents'],
                    'dispute_summary': '双方已到村委会调解室进行协商。',
                    'application_requests': ['恢复原状', '书面道歉'],
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(update_response.status_code, 200)
        updated_item = update_response.json()['item']
        self.assertEqual(updated_item['status'], '已调解')
        self.assertEqual(updated_item['occurrence_location'], '村委会调解室')

        export_response = self.client.get('/api/mediations/export/?status=已调解')
        self.assertEqual(export_response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            export_response['Content-Type'],
        )

        document_response = self.client.get(f'/api/mediations/{record_id}/application-document/')
        self.assertEqual(document_response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            document_response['Content-Type'],
        )
        decoded_disposition = ''.join(
            part.decode(charset or 'utf-8') if isinstance(part, bytes) else part
            for part, charset in decode_header(document_response['Content-Disposition'])
        )
        self.assertIn('attachment; filename=', decoded_disposition)
        self.assertIn('TJ-001.docx', decoded_disposition)

        delete_response = self.client.delete(f'/api/mediations/{record_id}/')
        self.assertEqual(delete_response.status_code, 200)
        self.assertFalse(MediationRecord.objects.filter(id=record_id).exists())

    def test_mediation_list_returns_actionable_message_when_tables_are_missing(self):
        with patch(
            'village_affairs.views.list_mediation_records',
            side_effect=ProgrammingError(
                '(1146, "Table \'country_manage_system.village_affairs_mediationrecord\' doesn\'t exist")'
            ),
        ):
            response = self.client.get('/api/mediations/')

        self.assertEqual(response.status_code, 503)
        payload = response.json()
        self.assertEqual(payload['error_code'], 'mediation_tables_missing')
        self.assertIn('manage.py migrate', payload['message'])

    def test_organization_member_create_query_update_delete(self):
        resident = Resident.objects.create(
            full_name='组织成员',
            gender='男',
            identity_number='110101199003030033',
            birth_date='1990-03-03',
            ethnicity='汉族',
            phone='13800003333',
            address='振兴路33号',
            political_status='中共党员',
            status='正常',
        )

        create_response = self.client.post(
            '/api/org-structure/create/',
            data=json.dumps(
                {
                    'resident_id': resident.id,
                    'org_type': '党(总)支部委员会',
                    'position': '支部书记',
                    'term_number': '第十二届',
                    'term_start': '2026-01-01',
                    'term_end': '2031-12-31',
                    'status': '现任',
                    'notes': '测试备注',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 201)
        item_id = create_response.json()['item']['id']

        list_response = self.client.get('/api/org-structure/?org_type=党(总)支部委员会&status=现任&keyword=支部')
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(list_payload['pagination']['total'], 1)
        self.assertEqual(list_payload['items'][0]['position'], '支部书记')
        self.assertEqual(list_payload['items'][0]['full_name'], '组织成员')

        update_response = self.client.put(
            f'/api/org-structure/{item_id}/',
            data=json.dumps(
                {
                    'position': '副书记',
                    'status': '历届',
                    'term_end': '2025-12-31',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.json()['item']['position'], '副书记')
        self.assertEqual(update_response.json()['item']['status'], '历届')

        export_response = self.client.get('/api/org-structure/export/?org_type=党(总)支部委员会')
        self.assertEqual(export_response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            export_response['Content-Type'],
        )

        delete_response = self.client.delete(f'/api/org-structure/{item_id}/')
        self.assertEqual(delete_response.status_code, 200)
        self.assertFalse(OrganizationMember.objects.filter(id=item_id).exists())

    def test_organization_member_bulk_delete(self):
        resident1 = Resident.objects.create(
            full_name='组织甲',
            identity_number='110101199404040044',
            status='正常',
        )
        resident2 = Resident.objects.create(
            full_name='组织乙',
            identity_number='110101199505050055',
            status='正常',
        )
        member1 = OrganizationMember.objects.create(
            resident=resident1,
            full_name=resident1.full_name,
            identity_number=resident1.identity_number,
            org_type='村居民委员会',
            position='委员',
            status='现任',
        )
        member2 = OrganizationMember.objects.create(
            resident=resident2,
            full_name=resident2.full_name,
            identity_number=resident2.identity_number,
            org_type='村居民委员会',
            position='委员',
            status='现任',
        )

        response = self.client.post(
            '/api/org-structure/bulk-delete/',
            data=json.dumps({'ids': [member1.id, member2.id]}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(OrganizationMember.objects.filter(id=member1.id).exists())
        self.assertFalse(OrganizationMember.objects.filter(id=member2.id).exists())

    def test_party_member_create_query_update_and_delete(self):
        resident = Resident.objects.create(
            full_name='党员甲',
            gender='男',
            identity_number='110101199112120011',
            birth_date='1991-12-12',
            ethnicity='汉族',
            phone='13800000088',
            address='先锋路1号',
            education_level='大专',
            political_status='中共党员',
            status='正常',
        )

        create_response = self.client.post(
            '/api/party-members/create/',
            data=json.dumps(
                {
                    'resident_id': resident.id,
                    'source': '居民档案',
                    'member_type': '中共党员',
                    'join_party_date': '2020-07-01',
                    'becoming_full_member_date': '2021-07-01',
                    'party_branch': '第一党支部',
                    'current_position': '组织委员',
                    'monthly_party_fee': '25.50',
                    'status': '正常',
                    'transfer_records': [
                        {
                            'transfer_type': '组织关系转接',
                            'transfer_date': '2024-01-01',
                            'from_branch': '原党支部',
                            'to_branch': '第一党支部',
                            'reason': '组织调整',
                        }
                    ],
                    'position_records': [
                        {
                            'branch_name': '第一党支部',
                            'position_name': '组织委员',
                            'start_date': '2024-01-01',
                            'is_current': True,
                        }
                    ],
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 201)
        item_id = create_response.json()['item']['id']

        list_response = self.client.get('/api/party-members/?full_name=党员甲&party_branch=第一党支部')
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(list_payload['pagination']['total'], 1)
        self.assertEqual(list_payload['items'][0]['member_type'], '中共党员')

        detail_response = self.client.get(f'/api/party-members/{item_id}/')
        self.assertEqual(detail_response.status_code, 200)
        detail_payload = detail_response.json()
        self.assertEqual(len(detail_payload['transfer_records']), 1)
        self.assertEqual(len(detail_payload['position_records']), 1)
        self.assertEqual(detail_payload['current_position'], '组织委员')

        update_response = self.client.put(
            f'/api/party-members/{item_id}/',
            data=json.dumps(
                {
                    'resident_id': resident.id,
                    'source': '居民档案',
                    'full_name': '党员甲',
                    'identity_number': resident.identity_number,
                    'gender': resident.gender,
                    'birth_date': '1991-12-12',
                    'ethnicity': resident.ethnicity,
                    'education_level': resident.education_level,
                    'phone': resident.phone,
                    'address': resident.address,
                    'member_type': '预备党员',
                    'join_party_date': '2020-07-01',
                    'becoming_full_member_date': '2021-07-01',
                    'party_branch': '第二党支部',
                    'current_position': '宣传委员',
                    'monthly_party_fee': '30.00',
                    'status': '正常',
                    'transfer_records': [
                        {
                            'id': detail_payload['transfer_records'][0]['id'],
                            'transfer_type': '转入',
                            'transfer_date': '2025-01-01',
                            'from_branch': '原党支部',
                            'to_branch': '第二党支部',
                            'reason': '工作调整',
                        }
                    ],
                    'position_records': [
                        {
                            'id': detail_payload['position_records'][0]['id'],
                            'branch_name': '第二党支部',
                            'position_name': '宣传委员',
                            'start_date': '2025-01-01',
                            'is_current': True,
                        }
                    ],
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.json()['item']['member_type'], '预备党员')
        self.assertEqual(update_response.json()['item']['party_branch'], '第二党支部')

        export_response = self.client.get('/api/party-members/export/?party_branch=第二党支部')
        self.assertEqual(export_response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            export_response['Content-Type'],
        )
        export_workbook = openpyxl.load_workbook(BytesIO(export_response.content))
        export_sheet = export_workbook.active
        self.assertEqual(export_sheet.max_row, 2)
        self.assertEqual(export_sheet.cell(row=2, column=1).value, '党员甲')
        self.assertEqual(export_sheet.cell(row=2, column=13).value, '第二党支部')

        delete_response = self.client.delete(f'/api/party-members/{item_id}/')
        self.assertEqual(delete_response.status_code, 200)
        self.assertFalse(PartyMember.objects.filter(id=item_id).exists())

    def test_party_fee_generate_update_and_mark_paid(self):
        member = PartyMember.objects.create(
            full_name='党员乙',
            identity_number='110101199202020022',
            member_type='中共党员',
            party_branch='第一党支部',
            monthly_party_fee='18.00',
            status='正常',
        )

        generate_response = self.client.post(
            '/api/party-fees/generate/',
            data=json.dumps({'fee_year': 2026, 'fee_month': 7, 'party_branch': '第一党支部'}),
            content_type='application/json',
        )
        self.assertEqual(generate_response.status_code, 200)
        self.assertEqual(PartyFeeRecord.objects.count(), 1)

        fee = PartyFeeRecord.objects.get(party_member=member, fee_year=2026, fee_month=7)
        self.assertEqual(str(fee.amount_due), '18.00')
        self.assertEqual(fee.payment_status, '待缴纳')

        list_response = self.client.get('/api/party-fees/?fee_year=2026&fee_month=7')
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(list_payload['pagination']['total'], 1)
        self.assertEqual(list_payload['stats']['pending_count'], 1)

        update_response = self.client.put(
            f'/api/party-fees/{fee.id}/',
            data=json.dumps(
                {
                    'amount_due': '20.00',
                    'amount_paid': '20.00',
                    'payment_status': '已缴纳',
                    'payment_date': '2026-07-20',
                    'notes': '现场收缴',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.json()['item']['payment_status'], '已缴纳')

        mark_response = self.client.post(f'/api/party-fees/{fee.id}/mark-paid/', data=json.dumps({}), content_type='application/json')
        self.assertEqual(mark_response.status_code, 200)

        export_response = self.client.get('/api/party-fees/export/?fee_year=2026&fee_month=7')
        self.assertEqual(export_response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            export_response['Content-Type'],
        )


class FarmlandFeatureTests(TestCase):
    def build_excel_file(self, headers, rows, filename='farmland.xlsx'):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return SimpleUploadedFile(
            filename,
            stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_farmland_import_list_summary_and_export(self):
        resident = Resident.objects.create(
            full_name='张承包',
            identity_number='110101199001010051',
            village_group='一组',
            status='正常',
        )

        upload_response = self.client.post(
            '/api/farmland/import/upload/',
            {
                'file': self.build_excel_file(
                    headers=['编号列', '村组列', '承包户列', '身份证列', '居民列', '位置列', '面积列', '状态列', '流转列', '确权列', '种植列', '变更列', '变更日期列', '备注列'],
                    rows=[
                        ['DK-001', '一组', '张承包', '110101199001010051', resident.id, '东山脚', '10.5', '正常', '已流转', '已确权', '水稻', '复垦整治', '2026-07-01', '首块地'],
                        ['DK-002', '一组', '张承包', '110101199001010051', resident.id, '西山坡', '5', '撂荒', '未流转', '未确权', '玉米', '待整治', '2026-07-02', '第二块地'],
                    ],
                )
            },
        )
        self.assertEqual(upload_response.status_code, 200)
        batch_id = upload_response.json()['batch_id']

        mapping = {
            'plot_code': '编号列',
            'village_group': '村组列',
            'contractor_name': '承包户列',
            'contractor_identity_number': '身份证列',
            'linked_resident_id': '居民列',
            'plot_location': '位置列',
            'area_mu': '面积列',
            'plot_status': '状态列',
            'transfer_status': '流转列',
            'confirmation_status': '确权列',
            'current_planting': '种植列',
            'latest_change': '变更列',
            'change_date': '变更日期列',
            'notes': '备注列',
        }

        preview_response = self.client.post(
            '/api/farmland/import/preview/',
            data=json.dumps({'batch_id': batch_id, 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(preview_response.status_code, 200)
        self.assertEqual(preview_response.json()['valid_rows'], 2)

        commit_response = self.client.post(
            '/api/farmland/import/commit/',
            data=json.dumps({'batch_id': batch_id, 'mapping': mapping}),
            content_type='application/json',
        )
        self.assertEqual(commit_response.status_code, 200)
        self.assertEqual(FarmlandRecord.objects.count(), 2)

        list_response = self.client.get('/api/farmland/?contractor_name=张承包')
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(list_payload['pagination']['total'], 2)
        self.assertEqual(list_payload['stats']['plot_count'], 2)
        self.assertEqual(list_payload['stats']['transferred_count'], 1)
        self.assertEqual(list_payload['stats']['abandoned_count'], 1)
        self.assertEqual(list_payload['stats']['reclaimed_count'], 1)
        self.assertEqual(list_payload['items'][0]['contractor_name'], '张承包')

        summary_response = self.client.get('/api/farmland/households/?village_group=一组')
        self.assertEqual(summary_response.status_code, 200)
        summary_payload = summary_response.json()
        self.assertEqual(summary_payload['pagination']['total'], 1)
        self.assertEqual(summary_payload['items'][0]['plot_count'], 2)
        self.assertEqual(summary_payload['items'][0]['transferred_count'], 1)
        self.assertEqual(summary_payload['items'][0]['confirmed_count'], 1)

        export_response = self.client.get('/api/farmland/export/?view=summary')
        self.assertEqual(export_response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            export_response['Content-Type'],
        )
        workbook = openpyxl.load_workbook(BytesIO(export_response.content))
        sheet = workbook.active
        self.assertEqual(sheet.title, '耕地按户汇总')
        self.assertEqual(sheet.cell(row=2, column=1).value, '张承包')
        self.assertEqual(sheet.cell(row=2, column=5).value, 2)

    def test_farmland_import_rejects_invalid_resident_reference(self):
        upload_response = self.client.post(
            '/api/farmland/import/upload/',
            {
                'file': self.build_excel_file(
                    headers=['编号列', '承包户列', '居民列'],
                    rows=[['DK-003', '李承包', 99999]],
                    filename='farmland-invalid.xlsx',
                )
            },
        )
        self.assertEqual(upload_response.status_code, 200)
        batch_id = upload_response.json()['batch_id']

        preview_response = self.client.post(
            '/api/farmland/import/preview/',
            data=json.dumps(
                {
                    'batch_id': batch_id,
                    'mapping': {
                        'plot_code': '编号列',
                        'contractor_name': '承包户列',
                        'linked_resident_id': '居民列',
                    },
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(preview_response.status_code, 200)
        payload = preview_response.json()
        self.assertEqual(payload['invalid_rows'], 1)
        self.assertIn('关联居民ID不存在', payload['errors'][0]['messages'])


class DatabaseBackupRestoreTests(TransactionTestCase):
    reset_sequences = True

    def test_database_backup_restore_overwrites_current_data(self):
        with TemporaryDirectory() as temp_dir, patch(
            'village_affairs.services.get_database_backup_dir',
            return_value=Path(temp_dir),
        ):
            resident = Resident.objects.create(
                full_name='原始居民',
                identity_number='110101199001010031',
                village_group='一组',
            )

            create_response = self.client.post(
                '/api/data-security/backups/create/',
                data=json.dumps({}),
                content_type='application/json',
            )
            self.assertEqual(create_response.status_code, 201)
            create_payload = create_response.json()
            backup_id = create_payload['item']['id']

            resident.full_name = '已被修改'
            resident.save(update_fields=['full_name'])
            Resident.objects.create(
                full_name='新增居民',
                identity_number='110101199001010032',
                village_group='二组',
            )

            restore_response = self.client.post(
                f'/api/data-security/backups/{backup_id}/restore/',
                data=json.dumps({}),
                content_type='application/json',
            )
            self.assertEqual(restore_response.status_code, 200)
            restore_payload = restore_response.json()
            self.assertIn('全量覆盖', restore_payload['message'])
            self.assertEqual(restore_payload['source_file_name'], create_payload['item']['file_name'])
            self.assertTrue(restore_payload['safety_backup_file_name'])

            resident.refresh_from_db()
            self.assertEqual(resident.full_name, '原始居民')
            self.assertFalse(Resident.objects.filter(identity_number='110101199001010032').exists())

    def test_database_backup_upload_restore_overwrites_current_data(self):
        with TemporaryDirectory() as temp_dir, patch(
            'village_affairs.services.get_database_backup_dir',
            return_value=Path(temp_dir),
        ):
            Resident.objects.create(
                full_name='上传恢复前数据',
                identity_number='110101199001010041',
                village_group='三组',
            )

            create_response = self.client.post(
                '/api/data-security/backups/create/',
                data=json.dumps({}),
                content_type='application/json',
            )
            self.assertEqual(create_response.status_code, 201)
            backup_file_path = Path(create_response.json()['item']['file_path'])
            self.assertTrue(backup_file_path.exists())

            Resident.objects.all().delete()
            Resident.objects.create(
                full_name='待覆盖数据',
                identity_number='110101199001010042',
                village_group='四组',
            )

            uploaded_file = SimpleUploadedFile(
                'restore-upload.json',
                backup_file_path.read_bytes(),
                content_type='application/json',
            )
            restore_response = self.client.post(
                '/api/data-security/backups/restore/upload/',
                {'file': uploaded_file},
            )
            self.assertEqual(restore_response.status_code, 200)
            restore_payload = restore_response.json()
            self.assertEqual(restore_payload['source_file_name'], 'restore-upload.json')
            self.assertTrue(restore_payload['safety_backup_file_name'])

            self.assertTrue(Resident.objects.filter(identity_number='110101199001010041').exists())
            self.assertFalse(Resident.objects.filter(identity_number='110101199001010042').exists())


class TodoReminderFeatureTests(TestCase):
    def test_todo_crud_bulk_read_and_summary(self):
        reminder_at = (timezone.now() - timedelta(minutes=10)).isoformat()
        create_response = self.client.post(
            '/api/todos/create/',
            data=json.dumps(
                {
                    'title': '走访老党员',
                    'content': '下午联系党员并完成走访记录',
                    'reminder_type': '任务',
                    'progress': '未开始',
                    'reminder_at': reminder_at,
                    'notes': '测试备注',
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 201)
        item_id = create_response.json()['item']['id']

        list_response = self.client.get('/api/todos/?view=未读')
        self.assertEqual(list_response.status_code, 200)
        list_payload = list_response.json()
        self.assertEqual(list_payload['pagination']['total'], 1)
        self.assertEqual(list_payload['stats']['due_count'], 1)
        self.assertEqual(list_payload['items'][0]['title'], '走访老党员')

        summary_response = self.client.get('/api/todos/summary/')
        self.assertEqual(summary_response.status_code, 200)
        self.assertEqual(summary_response.json()['due_count'], 1)
        self.assertEqual(summary_response.json()['not_started_count'], 1)
        self.assertEqual(summary_response.json()['in_progress_count'], 0)
        self.assertEqual(summary_response.json()['completed_count'], 0)

        bulk_read_response = self.client.post(
            '/api/todos/bulk-read/',
            data=json.dumps({'ids': [item_id], 'is_read': True}),
            content_type='application/json',
        )
        self.assertEqual(bulk_read_response.status_code, 200)
        self.assertTrue(TodoReminder.objects.get(id=item_id).is_read)
        refreshed_summary_response = self.client.get('/api/todos/summary/')
        self.assertEqual(refreshed_summary_response.status_code, 200)
        self.assertEqual(refreshed_summary_response.json()['due_count'], 0)

        update_response = self.client.put(
            f'/api/todos/{item_id}/',
            data=json.dumps(
                {
                    'title': '走访老党员',
                    'content': '下午联系党员并完成走访记录',
                    'reminder_type': '任务',
                    'progress': '已完成',
                    'reminder_at': reminder_at,
                    'notes': '已完成',
                    'is_read': True,
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.json()['item']['status'], '已完成')
        completed_summary_response = self.client.get('/api/todos/summary/')
        self.assertEqual(completed_summary_response.status_code, 200)
        self.assertEqual(completed_summary_response.json()['not_started_count'], 0)
        self.assertEqual(completed_summary_response.json()['completed_count'], 1)

        delete_response = self.client.delete(f'/api/todos/{item_id}/')
        self.assertEqual(delete_response.status_code, 200)
        self.assertFalse(TodoReminder.objects.filter(id=item_id).exists())

    def test_reminder_rules_cancel_birthday_and_keep_party_fee(self):
        birthday_rule = ReminderRule.objects.create(
            category=ReminderRule.CATEGORY_BIRTHDAY,
            rule_name='生日提醒',
            age_value=60,
            remind_days=7,
            is_enabled=True,
        )
        TodoReminder.objects.create(
            title='寿星提醒',
            reminder_type=TodoReminder.TYPE_BIRTHDAY,
            progress=TodoReminder.PROGRESS_NOT_STARTED,
            source_type=TodoReminder.SOURCE_BIRTHDAY_RULE,
            source_identifier=f'birthday:{birthday_rule.id}:1:{timezone.localdate().year}',
        )

        rules_response = self.client.get('/api/reminder-rules/')
        self.assertEqual(rules_response.status_code, 200)
        rules_payload = rules_response.json()
        self.assertEqual(rules_payload['birthday_rules'], [])
        self.assertGreaterEqual(len(rules_payload['party_fee_rules']), 1)
        self.assertFalse(ReminderRule.objects.filter(category=ReminderRule.CATEGORY_BIRTHDAY).exists())
        self.assertFalse(TodoReminder.objects.filter(reminder_type=TodoReminder.TYPE_BIRTHDAY).exists())

        create_rule_response = self.client.post(
            '/api/reminder-rules/create/',
            data=json.dumps(
                {
                    'category': 'party_fee',
                    'rule_name': '月末党费提醒',
                    'reminder_time': '08:30:00',
                    'is_month_end': True,
                    'is_enabled': True,
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(create_rule_response.status_code, 201)
        created_rule_id = create_rule_response.json()['item']['id']

        update_rule_response = self.client.put(
            f'/api/reminder-rules/{created_rule_id}/',
            data=json.dumps(
                {
                    'category': 'party_fee',
                    'rule_name': '月末党费提醒',
                    'reminder_time': '08:00:00',
                    'is_month_end': False,
                    'reminder_day': 28,
                    'is_enabled': False,
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(update_rule_response.status_code, 200)
        self.assertFalse(update_rule_response.json()['item']['is_enabled'])

        delete_rule_response = self.client.delete(f'/api/reminder-rules/{created_rule_id}/')
        self.assertEqual(delete_rule_response.status_code, 200)
        self.assertFalse(ReminderRule.objects.filter(id=created_rule_id).exists())

        rejected_birthday_response = self.client.post(
            '/api/reminder-rules/create/',
            data=json.dumps(
                {
                    'category': 'birthday',
                    'rule_name': '60岁生日提醒',
                    'age_value': 60,
                    'remind_days': 7,
                    'reminder_time': '09:00:00',
                    'is_enabled': True,
                }
            ),
            content_type='application/json',
        )
        self.assertEqual(rejected_birthday_response.status_code, 400)
        self.assertIn('生日提醒已取消', rejected_birthday_response.json()['message'])
